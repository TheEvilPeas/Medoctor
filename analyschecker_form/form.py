import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import webbrowser
def show_microsoft_login_dialog(flow: dict):
    verification_uri = cell_str(flow.get("verification_uri")) or "https://www.microsoft.com/link"
    user_code = cell_str(flow.get("user_code"))
    message = (
        "Чтобы получить доступ к таблице, нужно войти в Microsoft.\n\n"
        f"1. Открой страницу:\n{verification_uri}\n\n"
        f"2. Введи код:\n{user_code}\n\n"
        "3. Подтверди вход в браузере, затем вернись в программу."
    )

    print("\n=== ВХОД В MICROSOFT ДЛЯ ДОСТУПА К таблице ===")
    print(f"Открой страницу {verification_uri} и введи код {user_code}")
    print("Жду подтверждения входа...\n")

    root = tk._default_root
    temp_root = None

    try:
        if root is None:
            temp_root = tk.Tk()
            temp_root.withdraw()
            root = temp_root

        dialog = tk.Toplevel(root)
        dialog.title("Вход в Microsoft")
        dialog.resizable(False, False)
        dialog.transient(root)
        dialog.attributes('-topmost', True)
        dialog.grab_set()

        tk.Label(
            dialog,
            text="Вход в Microsoft для доступа к таблице",
            font=("Arial", 11, "bold"),
            anchor="w",
            justify="left",
        ).pack(fill="x", padx=14, pady=(14, 8))

        text_frame = tk.Frame(dialog)
        text_frame.pack(fill="both", expand=True, padx=14)

        message_label = tk.Label(
            text_frame,
            text=message,
            justify="left",
            anchor="w",
            wraplength=520,
        )
        message_label.pack(fill="x")

        code_frame = tk.Frame(dialog)
        code_frame.pack(fill="x", padx=14, pady=(12, 0))

        tk.Label(code_frame, text="Код:").pack(side="left")
        code_entry = tk.Entry(code_frame, width=20, justify="center")
        code_entry.pack(side="left", padx=(8, 8))
        code_entry.insert(0, user_code)
        code_entry.selection_range(0, tk.END)

        def copy_code():
            dialog.clipboard_clear()
            dialog.clipboard_append(user_code)
            dialog.update_idletasks()
            status_var.set("Код скопирован в буфер обмена")

        def open_link():
            try:
                webbrowser.open(verification_uri)
                status_var.set("Страница входа открыта в браузере")
            except Exception as e:
                status_var.set(f"Не удалось открыть браузер: {e}")

        buttons = tk.Frame(dialog)
        buttons.pack(fill="x", padx=14, pady=(14, 10))

        tk.Button(buttons, text="Открыть страницу", command=open_link).pack(side="left")
        tk.Button(buttons, text="Скопировать код", command=copy_code).pack(side="left", padx=(8, 0))
        tk.Button(buttons, text="Продолжить", command=dialog.destroy, bg="#4CAF50", fg="white").pack(side="right")

        status_var = tk.StringVar(value="После входа в браузере нажми 'Продолжить'")
        tk.Label(dialog, textvariable=status_var, fg="#555", anchor="w", justify="left").pack(fill="x", padx=14, pady=(0, 14))

        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        sw = dialog.winfo_screenwidth()
        sh = dialog.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        dialog.geometry(f"{w}x{h}+{x}+{y}")

        open_link()
        dialog.lift()
        dialog.focus_force()
        code_entry.focus_set()
        dialog.wait_window()
    finally:
        if temp_root is not None:
            try:
                temp_root.destroy()
            except Exception:
                pass
import re
import os
import json
import base64
from datetime import date, datetime
from tempfile import NamedTemporaryFile
from zipfile import BadZipFile
from urllib.parse import urlparse, parse_qs, unquote
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import xlrd
import requests
import xml.etree.ElementTree as ET

try:
    import msal
except ImportError:
    msal = None

MY_TABLE_DOWNLOAD_URL = "https://onedrive.live.com/:x:/g/personal/b353e867b45196ff/IQDKxFa4RLhbRobs6usMN0BSAS6bA1xFfl71nCnhAnUTh2w?e=mmcvXL&redeem=aHR0cHM6Ly8xZHJ2Lm1zL3gvYy9iMzUzZTg2N2I0NTE5NmZmL0lRREt4RmE0UkxoYlJvYnM2dXNNTjBCU0FTNmJBMXhGZmw3MW5DbmhBblVUaDJ3P2U9bW1jdlhM"
DEFAULT_MY_TABLE_DOWNLOAD_URL = MY_TABLE_DOWNLOAD_URL
AZURE_CLIENT_ID = "17b59301-c746-4e1a-89fa-942ef6465f00"                  # client id приложения Azure / Entra
AZURE_TENANT = "consumers"            # для личного OneDrive: consumers
TOKEN_CACHE_FILE = "ms_token_cache.json"
DETAILS_TABLE_FILE = "details_table_nov.xlsx"           # локальный файл details_table
DETAILS_TABLE_SHEET_NAME = "Прейскурант"      # имя листа в details_table
DETAILS_TABLE_START_ROW = 1            # номер строки, с которой начинаем добавлять записи; можно переопределить из главной программы
DETAILS_TABLE_END_ROW = None           # номер конечной строки включительно; None = читать до конца; можно переопределить из главной программы
PRICE_FILE = "input/price.xls"
PRICE_MAIN_SHEET_NAME = "1"
PRICE_EXTRA_SHEET_NAME = "2"
RESULT_FILE = "compare_result.xlsx"

DEFAULT_DETAILS_TABLE_FILE = DETAILS_TABLE_FILE
DEFAULT_RESULT_FILE = RESULT_FILE

MY_SHEET_NAME = "2025"
MY_TABLE_START_ROW = 1
MY_TABLE_END_ROW = None
DEFAULT_MY_SHEET_NAME = MY_SHEET_NAME
DEFAULT_MY_TABLE_START_ROW = MY_TABLE_START_ROW
DEFAULT_MY_TABLE_END_ROW = MY_TABLE_END_ROW
class AnalysCheckerForm(tk.Frame):
    def __init__(self, parent, main_app=None):
        super().__init__(parent)
        self.main_app = main_app
        self.notification_label = None

        self.my_table_url_var = tk.StringVar(value=MY_TABLE_DOWNLOAD_URL)
        self.my_sheet_name_var = tk.StringVar(value=MY_SHEET_NAME)
        self.my_start_row_var = tk.StringVar(value=str(MY_TABLE_START_ROW))
        self.my_end_row_var = tk.StringVar(value="" if MY_TABLE_END_ROW is None else str(MY_TABLE_END_ROW))
        self.details_file_var = tk.StringVar(value=DETAILS_TABLE_FILE)
        self.details_sheet_name_var = tk.StringVar(value=DETAILS_TABLE_SHEET_NAME)
        self.result_file_var = tk.StringVar(value=RESULT_FILE)
        self.status_var = tk.StringVar(value="Заполни параметры и нажми 'Запустить проверку'")

        self.build_ui()

    def build_ui(self):
        self.pack(fill="both", expand=True)
        self.columnconfigure(1, weight=1)

        row = 0
        tk.Label(self, text="Ссылка OneDrive на таблицу").grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        tk.Entry(self, textvariable=self.my_table_url_var).grid(row=row, column=1, sticky="ew", padx=10, pady=(10, 0))

        row += 1
        tk.Label(self, text="Название листа из моей таблицы ").grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        tk.Entry(self, textvariable=self.my_sheet_name_var).grid(row=row, column=1, sticky="ew", padx=10, pady=(10, 0))

        row += 1
        tk.Label(self, text="Строка начала проверки моей таблицы").grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        tk.Entry(self, textvariable=self.my_start_row_var).grid(row=row, column=1, sticky="ew", padx=10, pady=(10, 0))

        row += 1
        tk.Label(self, text="Строка конца проверки моей таблицы(пусто = до конца)").grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        tk.Entry(self, textvariable=self.my_end_row_var).grid(row=row, column=1, sticky="ew", padx=10, pady=(10, 0))

        row += 1
        tk.Label(self, text="Файл детализации").grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        details_row = tk.Frame(self)
        details_row.grid(row=row, column=1, sticky="ew", padx=10, pady=(10, 0))
        details_row.grid_columnconfigure(0, weight=1)
        tk.Entry(details_row, textvariable=self.details_file_var).grid(row=0, column=0, sticky="ew")
        tk.Button(details_row, text="Выбрать...", command=self.choose_details_file).grid(row=0, column=1, padx=(8, 0))

        row += 1
        tk.Label(self, text="Лист детализации").grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        tk.Entry(self, textvariable=self.details_sheet_name_var).grid(row=row, column=1, sticky="ew", padx=10, pady=(10, 0))

        row += 1
        tk.Label(self, text="Итоговый файл").grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        result_row = tk.Frame(self)
        result_row.grid(row=row, column=1, sticky="ew", padx=10, pady=(10, 0))
        result_row.grid_columnconfigure(0, weight=1)
        tk.Entry(result_row, textvariable=self.result_file_var).grid(row=0, column=0, sticky="ew")
        tk.Button(result_row, text="Сохранить как...", command=self.choose_result_file).grid(row=0, column=1, padx=(8, 0))

        row += 1
        tk.Label(self, textvariable=self.status_var, fg="#555", justify="left", anchor="w").grid(
            row=row, column=0, columnspan=2, sticky="ew", padx=10, pady=(15, 0)
        )

        row += 1
        button_row = tk.Frame(self)
        button_row.grid(row=row, column=0, columnspan=2, sticky="ew", padx=10, pady=15)
        tk.Button(button_row, text="Сбросить", command=self.reset_defaults).pack(side="left")
        tk.Button(button_row, text="Запустить проверку", command=self.run_check, bg="#4CAF50", fg="white").pack(side="right")
        tk.Button(button_row, text="Редактировать заключения", command=self.open_conclusions_editor).pack(side="right", padx=(0, 8))

        row += 1
        self.notification_label = tk.Label(
            self,
            text="",
            fg="white",
            bg="#333",
            bd=1,
            relief="solid",
            padx=10,
            pady=5,
        )
        self.notification_label.place_forget()

    def choose_details_file(self):
        path = filedialog.askopenfilename(
            parent=self.winfo_toplevel(),
            title="Выберите файл детализации",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if path:
            self.details_file_var.set(path)

    def choose_result_file(self):
        path = filedialog.asksaveasfilename(
            parent=self.winfo_toplevel(),
            title="Сохранить результат как",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=os.path.basename(self.result_file_var.get().strip() or DEFAULT_RESULT_FILE),
        )
        if path:
            self.result_file_var.set(path)

    def reset_defaults(self):
        self.my_table_url_var.set(DEFAULT_MY_TABLE_DOWNLOAD_URL)
        self.my_sheet_name_var.set(DEFAULT_MY_SHEET_NAME)
        self.my_start_row_var.set(str(DEFAULT_MY_TABLE_START_ROW))
        self.my_end_row_var.set("" if DEFAULT_MY_TABLE_END_ROW is None else str(DEFAULT_MY_TABLE_END_ROW))
        self.details_file_var.set(DEFAULT_DETAILS_TABLE_FILE)
        self.details_sheet_name_var.set(DETAILS_TABLE_SHEET_NAME)
        self.result_file_var.set(DEFAULT_RESULT_FILE)
        self.status_var.set("Параметры сброшены")

    def show_notification(self, text, duration=3000, x_offset=10, y_offset=10):
        if not self.notification_label:
            return
        self.notification_label.config(text=text)
        self.notification_label.update_idletasks()
        self.notification_label.place(
            relx=1.0,
            rely=1.0,
            anchor="se",
            x=-x_offset,
            y=-y_offset,
        )
        self.after(duration, self.notification_label.place_forget)

    def run_check(self):
        my_table_url = self.my_table_url_var.get().strip()
        my_sheet_name = self.my_sheet_name_var.get().strip()
        my_start_text = self.my_start_row_var.get().strip()
        my_end_text = self.my_end_row_var.get().strip()
        details_file = self.details_file_var.get().strip()
        details_sheet_name = self.details_sheet_name_var.get().strip()
        result_file = self.result_file_var.get().strip()

        if not my_table_url:
            messagebox.showerror("Ошибка", "Укажи ссылку OneDrive на my_table")
            return
        if not my_sheet_name:
            messagebox.showerror("Ошибка", "Укажи название листа из моей таблицы")
            return
        if not details_file:
            messagebox.showerror("Ошибка", "Укажи файл детализации")
            return
        if not os.path.exists(details_file):
            messagebox.showerror("Ошибка", f"Файл детализации не найден:\n{details_file}")
            return
        if not details_sheet_name:
            messagebox.showerror("Ошибка", "Укажи имя листа детализации")
            return
        if not result_file:
            messagebox.showerror("Ошибка", "Укажи имя итогового файла")
            return

        try:
            my_start_row = int(my_start_text) if my_start_text else 1
        except ValueError:
            messagebox.showerror("Ошибка", "MY_TABLE_START_ROW должен быть целым числом")
            return

        if my_start_row < 1:
            messagebox.showerror("Ошибка", "MY_TABLE_START_ROW должен быть не меньше 1")
            return

        if my_end_text == "":
            my_end_row = None
        else:
            try:
                my_end_row = int(my_end_text)
            except ValueError:
                messagebox.showerror("Ошибка", "MY_TABLE_END_ROW должен быть пустым или целым числом")
                return
            if my_end_row < my_start_row:
                messagebox.showerror("Ошибка", "MY_TABLE_END_ROW не может быть меньше MY_TABLE_START_ROW")
                return

        old_my_table_download_url = MY_TABLE_DOWNLOAD_URL
        old_my_sheet_name = MY_SHEET_NAME
        old_my_table_start_row = MY_TABLE_START_ROW
        old_my_table_end_row = MY_TABLE_END_ROW
        old_details_table_file = DETAILS_TABLE_FILE
        old_details_table_sheet_name = DETAILS_TABLE_SHEET_NAME
        old_details_table_start_row = DETAILS_TABLE_START_ROW
        old_details_table_end_row = DETAILS_TABLE_END_ROW
        old_result_file = RESULT_FILE

        try:
            globals()["MY_TABLE_DOWNLOAD_URL"] = my_table_url
            globals()["MY_SHEET_NAME"] = my_sheet_name
            globals()["MY_TABLE_START_ROW"] = my_start_row
            globals()["MY_TABLE_END_ROW"] = my_end_row
            globals()["DETAILS_TABLE_FILE"] = details_file
            globals()["DETAILS_TABLE_SHEET_NAME"] = details_sheet_name
            globals()["RESULT_FILE"] = result_file

            self.status_var.set("Проверка выполняется...")
            self.update_idletasks()
            main()
            self.status_var.set(f"Готово. Результат: {result_file}")
            self.show_notification(f"Готово: {result_file}")
            messagebox.showinfo("Готово", f"Проверка завершена.\nРезультат сохранён в:\n{result_file}")
        except Exception as e:
            self.status_var.set(f"Ошибка: {e}")
            messagebox.showerror("Ошибка", f"Не удалось выполнить проверку:\n{e}")
        finally:
            globals()["MY_TABLE_DOWNLOAD_URL"] = old_my_table_download_url
            globals()["MY_SHEET_NAME"] = old_my_sheet_name
            globals()["MY_TABLE_START_ROW"] = old_my_table_start_row
            globals()["MY_TABLE_END_ROW"] = old_my_table_end_row
            globals()["DETAILS_TABLE_FILE"] = old_details_table_file
            globals()["DETAILS_TABLE_SHEET_NAME"] = old_details_table_sheet_name
            globals()["DETAILS_TABLE_START_ROW"] = old_details_table_start_row
            globals()["DETAILS_TABLE_END_ROW"] = old_details_table_end_row
            globals()["RESULT_FILE"] = old_result_file


def open_analyschecker_standalone():
    root = tk.Tk()
    root.title("Проверка анализов")
    root.geometry("980x420")
    form = AnalysCheckerForm(root, main_app=None)
    form.pack(fill="both", expand=True)
    root.mainloop()

# Моя таблица
MY_COL_2 = 2          # во 2 столбце ищем "анализы" и ФИО
MY_DATE_COL = 3       # в строке с ФИО тут дата, она относится ко всем его анализам
MY_ANALYSIS_COL = 4   # колонка с анализом
MY_SUM_COL = 5        # колонка с суммой

# details_table
# DETAILS_DATE_COL = 2   # дата в 1 столбце, действует до следующей новой даты
# DETAILS_FIO_COL = 6
# DETAILS_ANALYSIS_COL = 10
# DETAILS_SUM_COL = 15

DETAILS_DATE_COL = 2
DETAILS_FIO_COL = 1
DETAILS_ANALYSIS_COL = 3
DETAILS_SUM_COL = 4


def resolve_my_table_source(url: str) -> str:
    url = cell_str(url)
    if url:
        return url
    return None


def normalize_onedrive_share_link(share_link: str) -> str:
    """
    Приводит OneDrive ссылку к стабильному виду для Graph /shares/.../driveItem/content.

    Если пришла длинная onedrive.live.com ссылка с redeem=...,
    пытается достать из redeem исходную 1drv.ms ссылку.
    """
    share_link = cell_str(share_link)
    if not share_link:
        return ""

    parsed = urlparse(share_link)
    host = (parsed.netloc or "").lower()

    if "onedrive.live.com" not in host:
        return share_link

    query = parse_qs(parsed.query)
    redeem_value = query.get("redeem", [""])[0]
    redeem_value = cell_str(redeem_value)
    if not redeem_value:
        return share_link

    try:
        if redeem_value.startswith("http://") or redeem_value.startswith("https://"):
            decoded = unquote(redeem_value)
            return decoded or share_link

        normalized = redeem_value.replace("-", "+").replace("_", "/")
        padding = "=" * ((4 - len(normalized) % 4) % 4)
        decoded = base64.b64decode(normalized + padding).decode("utf-8", errors="strict")
        decoded = unquote(decoded)

        if decoded.startswith("http://") or decoded.startswith("https://"):
            return decoded
    except Exception:
        pass

    return share_link


def create_graph_share_content_url(share_link: str) -> str:
    """
    Делает Graph URL вида /shares/{token}/driveItem/content из обычной share-ссылки.
    Перед кодированием приводит ссылку к стабильному виду:
    если пришла длинная onedrive.live.com ссылка с redeem=...,
    достаёт из неё исходную 1drv.ms ссылку.
    """
    share_link = normalize_onedrive_share_link(share_link)
    if not share_link:
        return ""

    data_bytes64 = base64.b64encode(share_link.encode("utf-8"))
    data_bytes64_string = (
        data_bytes64.decode("utf-8")
        .replace("/", "_")
        .replace("+", "-")
        .rstrip("=")
    )
    return f"https://graph.microsoft.com/v1.0/shares/u!{data_bytes64_string}/driveItem/content"

def load_token_cache():
    if not os.path.exists(TOKEN_CACHE_FILE):
        return None
    try:
        with open(TOKEN_CACHE_FILE, "r", encoding="utf-8") as f:
            cache_data = f.read()
        cache = msal.SerializableTokenCache()
        cache.deserialize(cache_data)
        return cache
    except Exception:
        return None


def save_token_cache(cache):
    try:
        if cache and cache.has_state_changed:
            with open(TOKEN_CACHE_FILE, "w", encoding="utf-8") as f:
                f.write(cache.serialize())
    except Exception:
        pass


def acquire_graph_access_token() -> str:
    """
    Получает delegated access token через device code flow и кеширует его локально.
    Подход основан на OAuth-аутентификации для Microsoft Graph.
    """
    if msal is None:
        raise RuntimeError(
            "Не установлен пакет msal. Установи его командой: pip install msal"
        )

    client_id = cell_str(AZURE_CLIENT_ID)
    tenant = cell_str(AZURE_TENANT) or "consumers"
    if not client_id:
        raise RuntimeError(
            "Не заполнен AZURE_CLIENT_ID. Создай app registration в Azure и укажи client id."
        )

    authority = f"https://login.microsoftonline.com/{tenant}"
    scopes = ["Files.Read"]
    # Для MSAL PublicClientApplication в device code flow нельзя передавать
    # reserved scopes вроде offline_access/openid/profile вручную.
    # Иначе MSAL падает с ошибкой "You cannot use any scope value that is reserved".

    cache = load_token_cache()
    if cache is None:
        cache = msal.SerializableTokenCache()

    app = msal.PublicClientApplication(
        client_id=client_id,
        authority=authority,
        token_cache=cache,
    )

    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(scopes, account=accounts[0])
        if result and result.get("access_token"):
            save_token_cache(cache)
            return result["access_token"]

    flow = app.initiate_device_flow(scopes=scopes)
    if "user_code" not in flow:
        raise RuntimeError(f"Не удалось начать device code flow: {flow}")

    show_microsoft_login_dialog(flow)

    result = app.acquire_token_by_device_flow(flow)
    save_token_cache(cache)

    if not result or not result.get("access_token"):
        raise RuntimeError(f"Не удалось получить Graph token: {result}")

    return result["access_token"]


def try_download_via_graph_share_link(share_link: str) -> tuple[bytes | None, str]:
    """
    Пытается скачать файл через Microsoft Graph /shares/.../driveItem/content.
    Скачивание делаем потоком, чтобы не зависать на response.content.
    """
    graph_url = create_graph_share_content_url(share_link)
    if not graph_url:
        return None, "empty graph share url"

    try:
        token = acquire_graph_access_token()
    except Exception as e:
        return None, f"auth error: {e}"

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        "User-Agent": "Mozilla/5.0",
    }

    try:
        with requests.get(
            graph_url,
            headers=headers,
            timeout=(20, 60),   # connect timeout, read timeout
            allow_redirects=True,
            stream=True,
        ) as response:
            if response.status_code >= 400:
                try:
                    error_text = response.text[:1000]
                except Exception:
                    error_text = "<no response text>"
                return None, f"graph HTTP {response.status_code}: {error_text}"

            chunks = []
            total_size = 0
            max_size = 100 * 1024 * 1024  # 100 MB защитный лимит

            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if not chunk:
                    continue
                chunks.append(chunk)
                total_size += len(chunk)

                if total_size > max_size:
                    return None, f"graph download too large: {total_size} bytes"

            data = b"".join(chunks)

    except requests.Timeout:
        return None, "graph request timeout while downloading file"
    except requests.RequestException as e:
        return None, f"graph request error: {e}"

    if not data.startswith(b"PK"):
        preview = data[:300].decode("utf-8", errors="replace")
        return None, f"graph downloaded not xlsx: {preview!r}"

    return data, ""


def download_my_table_to_temp(url: str) -> str:
    """
    Скачивает my_table.xlsx через Microsoft Graph shares endpoint
    с delegated user token сразу во временный файл.
    """
    url = cell_str(url)
    if not url:
        raise RuntimeError("Пустая MY_TABLE_DOWNLOAD_URL")

    parsed = urlparse(url)
    host = (parsed.netloc or "").lower()
    if "1drv.ms" not in host and "onedrive.live.com" not in host:
        raise RuntimeError(
            "Для текущего режима MY_TABLE_DOWNLOAD_URL должна быть обычной OneDrive share-ссылкой (1drv.ms или onedrive.live.com)."
        )

    graph_url = create_graph_share_content_url(url)
    if not graph_url:
        raise RuntimeError("Не удалось собрать Graph URL для скачивания my_table")

    token = acquire_graph_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        "User-Agent": "Mozilla/5.0",
    }

    temp_path = None
    try:
        with requests.get(
            graph_url,
            headers=headers,
            timeout=(20, 60),
            allow_redirects=True,
            stream=True,
        ) as response:
            if response.status_code >= 400:
                try:
                    error_text = response.text[:1000]
                except Exception:
                    error_text = "<no response text>"
                raise RuntimeError(
                    f"Не удалось скачать my_table через Microsoft Graph: graph HTTP {response.status_code}: {error_text}"
                )

            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                temp_path = tmp.name
                total_size = 0
                max_size = 100 * 1024 * 1024  # 100 MB

                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if not chunk:
                        continue
                    tmp.write(chunk)
                    total_size += len(chunk)

                    if total_size > max_size:
                        raise RuntimeError(f"Скачанный файл слишком большой: {total_size} bytes")

        with open(temp_path, "rb") as f:
            head = f.read(2)
        if head != b"PK":
            with open(temp_path, "rb") as f:
                preview = f.read(300).decode("utf-8", errors="replace")
            raise RuntimeError(f"Скачанный файл не является xlsx: {preview!r}")

        return temp_path

    except requests.Timeout:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        raise RuntimeError("Не удалось скачать my_table через Microsoft Graph: timeout while downloading file")

    except requests.RequestException as e:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        raise RuntimeError(f"Не удалось скачать my_table через Microsoft Graph: request error: {e}")

    except Exception:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        raise


# --- Helper: open my_table workbook for copy ---
def open_my_table_workbook_for_copy(source: str):
    """
    Открывает книгу my_table для копирования второго листа в результат.
    Возвращает tuple: (workbook, temp_path_or_none).
    Если source — URL, файл будет временно скачан повторно специально для копирования листа.
    """
    source = resolve_my_table_source(source)

    if source.startswith("http://") or source.startswith("https://"):
        temp_path = download_my_table_to_temp(source)
        wb = load_workbook(temp_path, data_only=True, read_only=True)
        return wb, temp_path

    wb = load_workbook(source, data_only=True, read_only=True)
    return wb, None




def get_online_excel_rows(source: str, worksheet_name: str, start_row: int = 1, end_row: int | None = None) -> list[tuple[int, list]]:
    """
    Читает строки либо из my_table по прямой ссылке на скачивание .xlsx,
    либо из локального fallback-файла.
    Возвращает список кортежей: (реальный_номер_строки, значения_строки).
    """
    source = resolve_my_table_source(source)
    start_row = max(1, int(start_row or 1))

    if source.startswith("http://") or source.startswith("https://"):
        temp_path = download_my_table_to_temp(source)
        try:
            wb = load_workbook(temp_path, data_only=True, read_only=True)
        except BadZipFile as e:
            raise RuntimeError(
                "Скачанный через Microsoft Graph файл не является настоящим .xlsx."
            ) from e
        finally:
            try:
                os.remove(temp_path)
            except OSError:
                pass
    else:
        wb = load_workbook(source, data_only=True, read_only=True)

    ws = wb[worksheet_name] if worksheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    # Берём ещё +1 столбец справа, потому что в скачанной my_table колонки
    # иногда сдвигаются, и сумма оказывается в 6-м столбце.
    max_needed_col = max(MY_COL_2, MY_DATE_COL, MY_ANALYSIS_COL, MY_SUM_COL + 1)
    rows = [list(row) for row in ws.iter_rows(min_col=1, max_col=max_needed_col, values_only=True)]

    if end_row is None:
        selected = rows[start_row - 1:]
    else:
        end_row = max(start_row, int(end_row))
        selected = rows[start_row - 1:end_row]

    result = []
    real_row_num = start_row
    for row in selected:
        result.append((real_row_num, list(row)))
        real_row_num += 1

    return result







def cell_str(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def collapse_single_letter_splits(text: str) -> str:
    """
    Чинит только явные разрывы внутри одного слова ФИО.

    Должно склеивать:
    - 'Вик тор' -> 'Виктор'
    - 'Ива нов' -> 'Иванов'
    - 'Алек сандр' -> 'Александр'
    - 'Андр еевна' -> 'Андреевна'

    Но не должно склеивать нормальные соседние слова:
    - 'Анна Егоровна'
    - 'Горбунов Илья'
    - 'Тагат Сабир Оглы'
    """
    parts = text.split()
    if not parts:
        return text

    repaired = []
    i = 0
    while i < len(parts):
        current = parts[i]

        if i + 1 < len(parts):
            nxt = parts[i + 1]
            merged = current + nxt

            # Склеиваем только явные разрывы внутри одного слова.
            should_merge = (
                4 <= len(merged) <= 18
                and (
                    # Вик + тор, Ива + нов
                    (len(current) <= 3 and len(nxt) <= 8)
                    # Алек + сандр, Андр + еевна
                    or (len(current) == 4 and len(nxt) <= 6)
                    # Очень короткий префикс + длинный хвост
                    or (len(current) <= 2 and len(nxt) <= 12)
                )
            )

            if should_merge:
                repaired.append(merged)
                i += 2
                continue

        repaired.append(current)
        i += 1

    return " ".join(repaired)


def repair_source_fio_text(text: str) -> str:
    """
    Чинит ФИО прямо на уровне исходной строки my_table ДО дальнейшего разбора.
    Исправляем только зону ФИО слева, хвосты с номерами не трогаем.

    Пример:
    - 'Агеев Вик тор Алек сандрович 1111№123456'
      -> 'АГЕЕВ ВИКТОР АЛЕКСАНДРОВИЧ 1111№123456'
    """
    raw = cell_str(text)
    if not raw:
        return ""

    normalized = normalize_spaces(raw).replace("Ё", "Е").replace("ё", "е")
    tokens = normalized.split()
    if not tokens:
        return normalized

    fio_tokens = []
    tail_tokens = []
    tail_started = False

    for token in tokens:
        if not tail_started and not token.isdigit() and "№" not in token and not re.search(r"\d", token):
            fio_tokens.append(token)
        else:
            tail_started = True
            tail_tokens.append(token)

    fixed_fio = collapse_single_letter_splits(" ".join(fio_tokens))

    if tail_tokens:
        return f"{fixed_fio} {' '.join(tail_tokens)}".strip()
    return fixed_fio.strip()

def repair_known_fio_typos(text: str) -> str:
    """
    Точечная правка известных опечаток в ФИО после базового восстановления пробелов.
    Правим только самые частые реальные случаи, чтобы не ломать нормальные ФИО.
    """
    value = cell_str(text)
    if not value:
        return ""

    result = normalize_spaces(value).upper()

    fixes = {
        "АЛЕ САНДР": "АЛЕКСАНДР",
        "АЛЕ САНДРОВИЧ": "АЛЕКСАНДРОВИЧ",
        "АЛЕ САНДРОВНА": "АЛЕКСАНДРОВНА",
        "АЛЕСАНДР": "АЛЕКСАНДР",
        " АЛЕКСАНДРОВИЧЯ": " АЛЕКСАНДРОВИЧ",
        " АЛЕКСАНДРОВНАЯ": " АЛЕКСАНДРОВНА",
    }

    for bad, good in fixes.items():
        result = result.replace(bad, good)

    return normalize_spaces(result)

def normalize_fio(text: str) -> str:
    text = cell_str(text)
    text = text.replace("Ё", "Е").replace("ё", "е")
    text = normalize_spaces(text)
    text = collapse_single_letter_splits(text)
    text = normalize_spaces(text)
    return text.upper()

def fio_parts(text: str) -> list[str]:
    normalized = normalize_fio(text)
    if not normalized:
        return []
    return normalized.split()


def fio_name_surname_key(text: str) -> str:
    parts = fio_parts(text)
    if len(parts) >= 2:
        # Формат Фамилия Имя [Отчество]
        return f"{parts[0]}|{parts[1]}"
    return ""


# --- New helper functions ---
def fio_first_last_key(text: str) -> str:
    parts = fio_parts(text)
    if len(parts) >= 2:
        # Формат Имя|Фамилия
        return f"{parts[1]}|{parts[0]}"
    return ""


def fio_first_last_analysis_key(fio: str, analysis) -> str:
    first_last = fio_first_last_key(fio)
    analysis_key = normalize_analysis(analysis)
    if first_last and analysis_key:
        return f"{first_last}|{analysis_key}"
    return ""

def analysis_matches_strict_or_mo(value, target_analysis) -> bool:
    """
    Совпадение для этапов с анализом в самом условии поиска:
    - либо тот же анализ
    - либо М/О

    Сравнение делаем в двух режимах:
    1) обычная normalize_analysis
    2) мягкая форма без разделителей, чтобы 467-А, 467А и 467 а считались одним анализом
    """
    if is_mo_marker(value):
        return True

    analysis_key = normalize_analysis(value)
    target_key = normalize_analysis(target_analysis)
    if analysis_key and target_key and analysis_key == target_key:
        return True

    value_loose = re.sub(r"[^A-ZА-Я0-9]", "", analysis_key)
    target_loose = re.sub(r"[^A-ZА-Я0-9]", "", target_key)

    return bool(value_loose and target_loose and value_loose == target_loose)


def fio_surname_key(text: str) -> str:
    parts = fio_parts(text)
    if parts:
        return parts[0]
    return ""


# --- New helper functions ---
def fio_name_surname_key_reversed(text: str) -> str:
    parts = fio_parts(text)
    if len(parts) >= 2:
        # Формат Имя+Фамилия для более мягкого поиска
        return f"{parts[1]}|{parts[0]}"
    return ""


def fio_surname_only_key(text: str) -> str:
    return fio_surname_key(text)


# --- Gender fallback surname normalization ---
def normalize_surname_for_gender_fallback(surname: str) -> str:
    """
    Нормализует фамилию для мягкого поиска с учётом частого различия по полу:
    ГОРБУНОВ <-> ГОРБУНОВА, ИВАНОВ <-> ИВАНОВА и т.п.
    Работает как fallback для поиска по фамилии.
    """
    surname = normalize_fio(surname)
    if not surname:
        return ""

    if surname.endswith("ОВА"):
        return surname[:-1]
    if surname.endswith("ЕВА"):
        return surname[:-1]
    if surname.endswith("ИНА"):
        return surname[:-1]
    if surname.endswith("ЫНА"):
        return surname[:-1]

    return surname


def fio_surname_gender_fallback_key(text: str) -> str:
    surname = fio_surname_only_key(text)
    return normalize_surname_for_gender_fallback(surname)


# --- New helper functions ---
def fio_name_surname_analysis_key(fio: str, analysis) -> str:
    name_surname = fio_name_surname_key(fio)
    analysis_key = normalize_analysis(analysis)
    if name_surname and analysis_key:
        return f"{name_surname}|{analysis_key}"
    return ""


def fio_surname_analysis_key(fio: str, analysis) -> str:
    surname = fio_surname_key(fio)
    analysis_key = normalize_analysis(analysis)
    if surname and analysis_key:
        return f"{surname}|{analysis_key}"
    return ""

def normalize_analysis(text) -> str:
    """
    Приводит код/название анализа к единому виду.

    Примеры совпадений:
    - 1515 == 1515.0
    - ОБС103 == ОБС 103
    - 457К1ПАТ-А == 457
    - 511ГИЭСПБ == ФГДС+ТЕСТ
    - PRS1-INV в details_table должно совпадать с PRS в my_table
    - 377С-УРО == 377
    - 377УРО == 377
    - 518СПБ == 518
    - 301УРО == 301
    - 302УРО == 302
    - 305УРО == 305
    - 308УРО == 308
    - 3090УРО == 3090
    - 343УРО == 343
    - 441-А == 441
    - 446-А == 446
    - 459-А == 459
    - 459 НОС == 459
    - 459 РОТ == 459
    - 467-А == 467
    - 159ЯГ == 159
    - 321СВ == 321
    - 11HOMA == 11НОМА
    - 1601ОСТ == 1601
    - 6802PH == 6802
    """
    if text is None:
        return ""

    if isinstance(text, float):
        if text.is_integer():
            value = str(int(text))
        else:
            value = str(text).replace(",", ".").strip().upper()
    elif isinstance(text, int):
        value = str(text)
    else:
        value = cell_str(text).replace("Ё", "Е").replace("ё", "е")
        value = normalize_spaces(value).upper()
        if re.fullmatch(r"\d+\.0+", value):
            value = value.split(".", 1)[0]

    # Латинскую H часто пишут как кириллическую Н, для HOMA считаем это одним и тем же.
    value = value.replace("HOMA", "НОМА")

    compact = re.sub(r"\s+", "", value)

    # Специальные соответствия между details_table и my_table / прайсом
    if compact == "511ГИЭСПБ":
        return "ФГДСТЕСТ"
    if compact in {"ФГДС+ТЕСТ", "ФГДСТЕСТ"}:
        return "ФГДСТЕСТ"

    # ОБС 103 -> ОБС103
    m_obs = re.match(r"^(ОБС)\s*(\d+)$", value)
    if m_obs:
        return f"{m_obs.group(1)}{m_obs.group(2)}"

    # PRS1-INV из details_table должно совпадать с PRS из my_table
    if compact in {"PRS1-INV", "PRS1INV", "PRS"}:
        return "PRS"

    # 511ГИЭСПБ из details_table должно совпадать с ФГДС+ТЕСТ из my_table
    if compact.startswith("511ГИЭСПБ"):
        return "ФГДСТЕСТ"

    # Коды, где основным считается стартовый числовой код.
    # Примеры:
    # 457К1ПАТ-А -> 457
    # 377С-УРО -> 377
    # 377УРО -> 377
    # 518СПБ -> 518
    # 301УРО -> 301
    # 302УРО -> 302
    # 305УРО -> 305
    # 308УРО -> 308
    # 3090УРО -> 3090
    # 343УРО -> 343
    # 441-А -> 441
    # 446-А -> 446
    # 459-А -> 459
    # 459 НОС -> 459
    # 459 РОТ -> 459
    # 467-А -> 467
    # 159ЯГ -> 159
    # 321СВ -> 321
    # 1601ОСТ -> 1601
    # 6802PH -> 6802
    m_numeric_prefix = re.match(r"^(\d+)([A-ZА-Я].*)?$", compact)
    if m_numeric_prefix:
        number_part = m_numeric_prefix.group(1)
        suffix_part = m_numeric_prefix.group(2) or ""

        if suffix_part in {
            "",
            "-А",
            "А",
            "С-УРО",
            "УРО",
            "СПБ",
            "ОСТ",
            "PH",
            "ЯГ",
            "СВ",
            "НОС",
            "РОТ",
        } or compact.startswith("457"):
            return number_part

    return compact

def analysis_candidate_keys(text) -> set[str]:
    """
    Возвращает набор ключей для мягкого сопоставления анализа.

    Идея:
    - сохраняем строгий нормализованный ключ
    - если в анализе есть ведущий числовой код, добавляем его как мягкий ключ
    - если есть только цифры внутри, тоже добавляем слитный цифровой ключ

    Примеры:
    - 457К1ПАТ-А -> {"457"}
    - 459 НОС -> {"459"}
    - ОБС 103 -> {"ОБС103", "103"}
    - PRS1-INV -> {"PRS", "1"}
    - 11НОМА -> {"11НОМА", "11"}
    """
    key = normalize_analysis(text)
    if not key:
        return set()

    result = {key}

    m_prefix = re.match(r"^(\d+)", key)
    if m_prefix:
        result.add(m_prefix.group(1))

    digits = "".join(re.findall(r"\d+", key))
    if digits:
        result.add(digits)

    return {item for item in result if item}

def analysis_keys_match(left, right) -> bool:
    """
    Универсальное сравнение анализов:
    - сначала строгая нормализация
    - потом мягкое совпадение по числовым ключам

    Это позволяет находить совпадения не только по точному тексту,
    но и по базовому цифровому коду анализа.
    """
    left_keys = analysis_candidate_keys(left)
    right_keys = analysis_candidate_keys(right)

    if not left_keys or not right_keys:
        return False

    return bool(left_keys & right_keys)

def _strip_tail_tokens_for_fio(text: str) -> list[str]:
    """
    Берёт только ту часть строки, где потенциально находится ФИО,
    и отрезает хвосты с номерами, кодами, датами и прочим служебным мусором.
    """
    text = normalize_fio(text)
    tokens = text.split()
    result = []

    for token in tokens:
        if token.isdigit() or "№" in token:
            break
        if re.search(r"\d", token):
            break
        result.append(token)

    return result


def _repair_split_fio_tokens(tokens: list[str]) -> list[str]:
    """
    Восстанавливает ФИО из токенов, даже если пробелы попали внутрь
    нескольких слов сразу.

    Примеры:
    - ['ИВА', 'НОВ', 'ИВ', 'АН', 'ИВА', 'НОВИЧ'] -> ['ИВАНОВ', 'ИВАН', 'ИВАНОВИЧ']
    - ['ВИК', 'ТОР', 'СЕР', 'ГЕЕВИЧ'] -> ['ВИКТОР', 'СЕРГЕЕВИЧ']
    """
    if not tokens:
        return []

    tokens = tokens[:9]

    def build_groups(parts: list[str], group_count: int):
        results = []

        def rec(start: int, groups_left: int, acc: list[str]):
            if groups_left == 0:
                if start == len(parts):
                    results.append(acc[:])
                return

            remaining = len(parts) - start
            if remaining < groups_left:
                return

            # В одно слово можно склеить от 1 до 4 соседних кусков.
            max_take = min(4, remaining - groups_left + 1)
            for take in range(1, max_take + 1):
                word = "".join(parts[start:start + take])
                if 2 <= len(word) <= 16:
                    acc.append(word)
                    rec(start + take, groups_left - 1, acc)
                    acc.pop()

        rec(0, group_count, [])
        return results

    def score_grouping(words: list[str]) -> int:
        # Предпочитаем 3 слова, но допускаем 2.
        score = 0
        if len(words) == 2:
            score += 5

        # Штрафуем слишком короткие и слишком длинные слова.
        for w in words:
            if len(w) < 3:
                score += 20
            score += abs(len(w) - 7)

        # Предпочитаем, чтобы отчество было длиннее имени.
        if len(words) >= 3 and len(words[2]) < len(words[1]):
            score += 3

        return score

    candidates = []
    for group_count in (3, 2):
        candidates.extend(build_groups(tokens, group_count))

    if not candidates:
        return tokens[:3]

    best = min(candidates, key=score_grouping)
    return best

def extract_my_fio(text: str) -> str:
    # Сначала отрезаем хвост с номерами/кодами.
    tokens = _strip_tail_tokens_for_fio(text)
    if not tokens:
        return ""

    # 1) Пытаемся быстро починить попарные разрывы внутри слов.
    repaired_text = collapse_single_letter_splits(" ".join(tokens))
    repaired_tokens = repaired_text.split()

    # Спецслучай для ФИО с добавкой типа "ОГЛЫ" / "КЫЗЫ".
    if len(repaired_tokens) >= 4 and repaired_tokens[3] in {"ОГЛЫ", "КЫЗЫ"}:
        candidate = " ".join(repaired_tokens[:4])
        if re.fullmatch(r"[А-ЯA-Z-]+ [А-ЯA-Z-]+ [А-ЯA-Z-]+ (ОГЛЫ|КЫЗЫ)", candidate):
            return candidate

    if len(repaired_tokens) >= 3:
        candidate = " ".join(repaired_tokens[:3])
        if re.fullmatch(r"[А-ЯA-Z-]+ [А-ЯA-Z-]+ [А-ЯA-Z-]+", candidate):
            return candidate
    if len(repaired_tokens) == 2:
        candidate = " ".join(repaired_tokens)
        if re.fullmatch(r"[А-ЯA-Z-]+ [А-ЯA-Z-]+", candidate):
            return candidate

    # 2) Если не вышло, используем более тяжёлое восстановление.
    repaired = _repair_split_fio_tokens(tokens)
    if len(repaired) >= 4 and repaired[3] in {"ОГЛЫ", "КЫЗЫ"}:
        return " ".join(repaired[:4])
    if len(repaired) >= 3:
        return " ".join(repaired[:3])
    if len(repaired) == 2:
        return " ".join(repaired)

    return ""


def extract_my_fio_loose(text: str) -> str:
    """
    Более мягкое извлечение ФИО для случаев, когда внутри слов есть лишние пробелы,
    например:
    - 'Ива нов Вик тор Ива нович 1104№123456'
    - 'Ива нов Ив ан Ива нович'
    - 'Ива нов ив ан иван ович'
    """
    tokens = _strip_tail_tokens_for_fio(text)
    if not tokens:
        return ""

    repaired = _repair_split_fio_tokens(tokens)
    if len(repaired) >= 3:
        return " ".join(repaired[:3])
    if len(repaired) == 2:
        return " ".join(repaired)
    return ""


def looks_like_fio(text: str) -> bool:
    candidate = extract_my_fio(text)
    if not candidate:
        return False

    parts = candidate.split()

    # Допускаем обычные 2-3 слова и спецслучай из 4 слов с "ОГЛЫ" / "КЫЗЫ".
    if len(parts) == 4 and parts[3] in {"ОГЛЫ", "КЫЗЫ"}:
        return all(2 <= len(p) <= 16 for p in parts[:3])

    if not (2 <= len(parts) <= 3):
        return False

    return all(2 <= len(p) <= 16 for p in parts)


def contains_analyses_word(text: str) -> bool:
    return "анализы" in cell_str(text).lower()


def extract_google_fio(text: str) -> str:
    text = normalize_fio(text)

    m = re.match(
        r"^([А-ЯЁA-Z][А-ЯЁа-яёa-z-]+ [А-ЯЁA-Z][А-ЯЁа-яёa-z-]+(?: [А-ЯЁA-Z][А-ЯЁа-яёa-z-]+)?)",
        text,
    )
    if m:
        return m.group(1)
    return ""


def parse_date(value):
    if value is None:
        return None

    # Настоящие Excel-даты / datetime
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")

    text = cell_str(value)
    if not text:
        return None

    # Убираем время, если оно прицепилось к текстовой дате
    if " " in text:
        text = text.split(" ", 1)[0]

    # Разрешаем только реальные даты, а не коды вида 310583Ж5ЩМИ
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text):
        return text

    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        yyyy, mm, dd = text.split("-")
        return f"{dd}.{mm}.{yyyy}"

    return None

def format_output_date(value):
    """
    Приводит дату к виду дд.мм.гг для записи в результат.
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%d.%m.%y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%y")

    text = cell_str(value)
    if not text:
        return None

    if re.fullmatch(r"\d{2}\.\d{2}\.\d{2}", text):
        return text

    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text):
        dd, mm, yyyy = text.split(".")
        return f"{dd}.{mm}.{yyyy[-2:]}"

    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})(?:\s+.*)?$", text)
    if m:
        yyyy, mm, dd = m.groups()
        return f"{dd}.{mm}.{yyyy[-2:]}"

    parsed = parse_date(text)
    if parsed and re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", parsed):
        dd, mm, yyyy = parsed.split(".")
        return f"{dd}.{mm}.{yyyy[-2:]}"

    return text

def parse_money(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value

    text = cell_str(value).replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return text

def looks_like_date_value(value) -> bool:
    """
    Проверяет, похожа ли ячейка на дату.
    Нужна для подстройки под сдвинутые колонки в my_table.
    """
    if value is None:
        return False
    if isinstance(value, (datetime, date)):
        return True
    return parse_date(value) is not None


# --- Дополнительные функции для фильтрации значений суммы ---
def is_mo_marker(value) -> bool:
    """
    Разрешённые текстовые значения в столбце суммы, которые нужно учитывать при поиске.
    Принимаем только варианты М/О, М\О и МО.
    """
    text = cell_str(value).replace(" ", "").upper()
    return text in {"М/О", "М\\О", "МО"}


def is_matchable_sum_value(value) -> bool:
    """
    Для поиска учитываем:
    - любые числовые суммы
    - специальные текстовые пометки М/О, М\О, МО
    Всё остальное из столбца суммы для поиска игнорируем.
    """
    return isinstance(value, (int, float)) or is_mo_marker(value)


# --- Helper: фильтрация записей по анализу для этапов без анализа в ключе ---
def is_matchable_analysis_value(value, target_analysis) -> bool:
    """
    Для подстановки оставляем только записи, где в столбце анализа:
    - либо тот же анализ, что ищем
    - либо пометка М/О, М\О, МО
    Всё остальное не подставляем.
    """
    if is_mo_marker(value):
        return True

    analysis_key = normalize_analysis(value)
    target_key = normalize_analysis(target_analysis)

    if not analysis_key or not target_key:
        return False

    return analysis_key == target_key

def is_real_analysis_value(value) -> bool:
    """
    Оставляем только реальные анализы из столбца анализа.
    Нас интересуют:
    - коды/анализы, где есть цифры: 467-А, 455, ОБС90, 457К1ПАТ-А и т.п.
    - пометки М/О, М\О, МО

    Всё остальное вроде 'первич', 'экг', 'терапевт' и т.п.
    не считаем анализом и не используем для подстановки.
    """
    if is_mo_marker(value):
        return True

    analysis_key = normalize_analysis(value)
    if not analysis_key:
        return False

    return bool(re.search(r"\d", analysis_key))

def debug_my_sheet(filename, sheet_name, max_rows=20, max_cols=8):
    wb = load_workbook(filename, data_only=True, read_only=True)
    print("МОЯ ТАБЛИЦА: листы =", wb.sheetnames)

    if sheet_name not in wb.sheetnames:
        print(f"Лист '{sheet_name}' не найден")
        return

    ws = wb[sheet_name]
    print(f"Активный лист: {sheet_name}")

    print("\nПервые строки моей таблицы:")
    for row_index, row in enumerate(
        ws.iter_rows(min_col=1, max_col=max_cols, values_only=True), start=1
    ):
        if row_index > max_rows:
            break
        values = [cell_str(row[col]) if col < len(row) else "" for col in range(max_cols)]
        print(f"{row_index:>4}: {values}")
        col2 = values[1] if len(values) > 1 else ""
        if col2:
            print(f"      -> collapse_single_letter_splits: {collapse_single_letter_splits(col2)}")
            print(f"      -> extracted_fio: {extract_my_fio(col2)} | looks_like_fio={looks_like_fio(col2)}")


def parse_my_table(source, sheet_name, start_row=1, end_row=None):
    rows = get_online_excel_rows(source, sheet_name, start_row=start_row, end_row=end_row)

    records = []
    current_fio = None
    current_date = None

    for row_index, row in rows:
        if row_index % 5000 == 0:
            print(f"Обработка my_table: строка {row_index}")

        max_needed_col = max(MY_COL_2, MY_DATE_COL, MY_ANALYSIS_COL, MY_SUM_COL + 1)
        if len(row) < max_needed_col:
            row = list(row) + [None] * (max_needed_col - len(row))

        original_col2 = cell_str(row[MY_COL_2 - 1])
        col2 = repair_source_fio_text(original_col2)
        col2 = repair_known_fio_typos(col2)

        # Базовая схема my_table:
        # col 2 = ФИО, col 3 = дата, col 4 = анализ, col 5 = сумма.
        raw_date_cell = row[MY_DATE_COL - 1]
        raw_analysis_cell = row[MY_ANALYSIS_COL - 1]
        raw_sum_cell = row[MY_SUM_COL - 1]

        row_date = parse_date(raw_date_cell)
        analysis = cell_str(raw_analysis_cell)
        amount = parse_money(raw_sum_cell)

        # Иногда в скачанной my_table колонки сдвигаются вправо на 1:
        # col 3 пусто, col 4 = дата, col 5 = анализ, col 6 = сумма.
        # Тогда старый разбор даёт:
        # analysis = дата, sum = код анализа.
        # Исправляем это автоматически.
        shifted_date = raw_analysis_cell
        shifted_analysis = raw_sum_cell
        shifted_sum = row[MY_SUM_COL] if len(row) > MY_SUM_COL else None

        if (
            row_date is None
            and looks_like_date_value(shifted_date)
            and cell_str(shifted_analysis) != ""
        ):
            row_date = parse_date(shifted_date)
            analysis = cell_str(shifted_analysis)
            amount = parse_money(shifted_sum)

        # Продолжение того же человека на следующей строке:
        # ФИО уже нет, дата тоже часто пустая, но анализ и сумма стоят
        # в сдвинутых колонках (col 5 и col 6).
        # Пример:
        #   САЛАЗАНОВА ... | 1 | 2026-01-28 | 441-а | 630
        #   <пусто>        |   |            | 455   | 1320
        # Для таких строк сохраняем предыдущий current_fio/current_date
        # и читаем анализ из col 5, сумму из col 6.
        if (
            current_fio
            and row_date is None
            and cell_str(raw_analysis_cell) == ""
            and cell_str(raw_sum_cell) != ""
        ):
            shifted_cont_analysis = cell_str(raw_sum_cell)
            shifted_cont_sum = row[MY_SUM_COL] if len(row) > MY_SUM_COL else None

            if shifted_cont_analysis:
                analysis = shifted_cont_analysis
                amount = parse_money(shifted_cont_sum)


        # Если в строке есть ФИО, запоминаем текущего человека и его дату.
        if looks_like_fio(col2):
            current_fio = extract_my_fio(col2)
            current_date = row_date

        # Записываем вообще все записи, где уже известен человек и есть анализ.
        # Если ФИО пустое, но current_fio уже есть, значит это продолжение этого же человека.
        if current_fio and analysis:
            records.append(
                {
                    "date": current_date,
                    "fio": current_fio,
                    "analysis": analysis,
                    "sum": amount,
                    "row": row_index,
                }
            )

    matchable_records = sum(1 for item in records if is_matchable_sum_value(item.get("sum")))
    print(f"my_table parsed: всего записей={len(records)}, пригодных для поиска по сумме/М/О={matchable_records}")
    return records

def find_details_start_row(rows, max_scan_rows: int = 5000) -> int:
    """
    Ищет первую строку детализации, где в 1-м столбце стоит число 1.
    rows — список строк, где каждая строка уже представлена списком значений.
    Если не найдено, возвращает 1.
    """
    scan_limit = min(max_scan_rows, len(rows))

    for row_index in range(scan_limit):
        row = rows[row_index] if row_index < len(rows) else []
        value = row[0] if row else None

        if value == 1:
            return row_index + 1

        text = cell_str(value)
        if text == "1":
            return row_index + 1

        if isinstance(value, float) and value.is_integer() and int(value) == 1:
            return row_index + 1

    return 1

def parse_details_table(filename, worksheet_name, start_row=1, end_row=None):
    ext = os.path.splitext(filename)[1].lower()
    max_needed_col = max(DETAILS_DATE_COL, DETAILS_FIO_COL, DETAILS_ANALYSIS_COL, DETAILS_SUM_COL)

    rows = []

    if ext == ".xls":
        book = xlrd.open_workbook(filename, formatting_info=False)
        try:
            sheet = book.sheet_by_name(worksheet_name)
        except xlrd.biffh.XLRDError:
            sheet = book.sheet_by_index(0)

        for row_index in range(sheet.nrows):
            row_len = sheet.row_len(row_index)
            row = []
            for col_index in range(max_needed_col):
                if col_index < row_len:
                    row.append(sheet.cell_value(row_index, col_index))
                else:
                    row.append(None)
            rows.append(row)
    else:
        wb = load_workbook(filename, data_only=True, read_only=True)
        ws = wb[worksheet_name] if worksheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        for row in ws.iter_rows(min_col=1, max_col=max_needed_col, values_only=True):
            row_values = list(row)
            if len(row_values) < max_needed_col:
                row_values += [None] * (max_needed_col - len(row_values))
            rows.append(row_values)

    detected_start_row = find_details_start_row(rows)
    if start_row is None or int(start_row or 0) <= 1:
        start_row = detected_start_row

    records = []
    pending_date = None
    current_date = None

    for row_index, row in enumerate(rows, start=1):
        if row_index < start_row:
            continue
        if end_row is not None and row_index > end_row:
            break

        if row_index % 5000 == 0:
            print(f"Обработка details_table: строка {row_index}")

        if len(row) < max_needed_col:
            row = list(row) + [None] * (max_needed_col - len(row))

        date_cell = parse_date(row[DETAILS_DATE_COL - 1])
        if date_cell:
            pending_date = date_cell

        raw_fio = cell_str(row[DETAILS_FIO_COL - 1])
        analysis = cell_str(row[DETAILS_ANALYSIS_COL - 1])
        amount = parse_money(row[DETAILS_SUM_COL - 1])

        if not raw_fio:
            continue

        fio = extract_google_fio(raw_fio)
        if not fio:
            continue

        if pending_date:
            current_date = pending_date
            pending_date = None

        if not analysis and amount is None:
            continue

        records.append(
            {
                "date": current_date,
                "fio": fio,
                "analysis": analysis,
                "sum": amount,
                "row": row_index,
            }
        )

    return records


def parse_price_table(filename, sheet_name):
    book = xlrd.open_workbook(filename, formatting_info=False)

    # Для .xls сначала пробуем лист по имени, иначе берём первый.
    try:
        sheet = book.sheet_by_name(sheet_name)
        print(f"Прайс: читаю лист '{sheet_name}'")
    except xlrd.biffh.XLRDError:
        sheet = book.sheet_by_index(0)
        print(f"Прайс: лист '{sheet_name}' не найден, читаю первый лист '{sheet.name}'")

    price_map = {}

    for row_index in range(9, sheet.nrows):  # Excel row 10 -> index 9
        # В прайсе название анализа находится во 2 столбце, цена — в 8 столбце.
        analysis_raw = cell_str(sheet.cell_value(row_index, 1))
        price_raw = sheet.cell_value(row_index, 7)

        if not analysis_raw:
            continue

        analysis_key = normalize_analysis(analysis_raw)
        if not analysis_key:
            continue

        price_value = parse_money(price_raw)

        # Если в цене ничего нет, такую строку пропускаем
        if price_value in (None, ""):
            continue

        price_map[analysis_key] = {
            "analysis": analysis_raw,
            "price": price_value,
            "row": row_index + 1,
        }

    return price_map


def parse_price2_table(filename, sheet_name):
    book = xlrd.open_workbook(filename, formatting_info=False)

    try:
        sheet = book.sheet_by_name(sheet_name)
        print(f"Прайс лист 2: читаю лист '{sheet_name}'")
    except xlrd.biffh.XLRDError:
        if book.nsheets > 1:
            sheet = book.sheet_by_index(1)
        else:
            sheet = book.sheet_by_index(0)
        print(f"Прайс лист 2: лист '{sheet_name}' не найден, читаю лист '{sheet.name}'")

    price_map = {}

    for row_index in range(9, sheet.nrows):  # Excel row 10 -> index 9
        analysis_raw = cell_str(sheet.cell_value(row_index, 1))
        price_raw = sheet.cell_value(row_index, 7)

        if not analysis_raw:
            continue

        analysis_key = normalize_analysis(analysis_raw)
        if not analysis_key:
            continue

        price_value = parse_money(price_raw)
        if price_value in (None, ""):
            continue

        price_map[analysis_key] = {
            "analysis": analysis_raw,
            "price": price_value,
            "row": row_index + 1,
        }

    return price_map

def print_records(title, records):
    print("\n" + "=" * 120)
    print(title)
    print("=" * 120)

    if not records:
        print("Ничего не найдено")
        return

    for i, item in enumerate(records, start=1):
        print(
            f"{i:>4}. [{item['row']}] "
            f"{item['date'] or '-'} | "
            f"{item['fio']} | "
            f"{item['analysis'] or '-'} | "
            f"{item['sum'] if item['sum'] is not None else '-'}"
        )

    print(f"\nИтого: {len(records)}")


def build_result_file(details_records, my_records, price_map, price2_map, my_table_source, my_sheet_name, result_file):
    # Готовим my_table для нового простого алгоритма:
    # для каждой записи сохраняем только нормализованные ФИО и анализ.
    prepared_my_records = []

    for item in my_records:
        fio_norm = normalize_fio(item.get("fio"))
        fio_parts_list = fio_parts(item.get("fio"))
        surname_name_key = ""
        surname_key = ""
        if len(fio_parts_list) >= 2:
            surname_name_key = f"{fio_parts_list[0]}|{fio_parts_list[1]}"
            surname_key = fio_parts_list[0]

        prepared_my_records.append(
            {
                "item": item,
                "fio_norm": fio_norm,
                "analysis_norm": normalize_analysis(item.get("analysis")),
                "surname_name_key": surname_name_key,
                "surname_key": surname_key,
                "surname_gender_key": normalize_surname_for_gender_fallback(surname_key),
            }
        )

    print(f"Подготовка my_table завершена: всего записей={len(prepared_my_records)}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Сравнение"

    copied_sheet_name = "Моя таблица"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    bold_font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    # Заголовки
    base_headers = ["ФИО details_table", "Дата details_table", "Анализ details_table", "Сумма details_table", "Цена по прайсу"]
    for col, header in enumerate(base_headers, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.fill = header_fill
        c.font = bold_font

    row_num = 2
    max_match_cols = 0

    for g in details_records:
        fio = g["fio"]
        date_val = g["date"]
        analysis = g["analysis"]
        google_sum = g["sum"]

        ws.cell(row=row_num, column=1, value=fio)
        ws.cell(row=row_num, column=2, value=format_output_date(date_val))
        ws.cell(row=row_num, column=3, value=cell_str(analysis))
        ws.cell(row=row_num, column=4, value=google_sum)

        normalized_analysis = normalize_analysis(analysis)

        # Для цены из прайса используем только строгое совпадение ключа.
        # Сначала ищем на листе 1, и только если там не нашли — смотрим лист 2.
        # Мягкий поиск по цифровым кандидатам здесь запрещён,
        # иначе 505Б может ошибочно подставить цену от 505.
        price_info = price_map.get(normalized_analysis)
        price_source = "price.xls"

        if price_info is None:
            price_info = price2_map.get(normalized_analysis)
            if price_info is not None:
                price_source = "price2.xlsx"


        price_value = price_info["price"] if price_info else None
        price_cell = ws.cell(row=row_num, column=5, value=price_value)

        if (
            isinstance(google_sum, (int, float))
            and isinstance(price_value, (int, float))
            and google_sum > price_value
        ):
            price_cell.fill = red_fill

        full_fio_key = normalize_fio(fio)
        normalized_analysis = normalize_analysis(analysis)
        details_parts_list = fio_parts(fio)
        details_surname_name_key = ""
        details_surname_key = ""
        details_surname_gender_key = ""
        if len(details_parts_list) >= 2:
            details_surname_name_key = f"{details_parts_list[0]}|{details_parts_list[1]}"
            details_surname_key = details_parts_list[0]
            details_surname_gender_key = normalize_surname_for_gender_fallback(details_surname_key)

        key = (full_fio_key, normalized_analysis)

        # Новый алгоритм по этапам:
        # 1) ФИО + анализ
        # 2) Фамилия + Имя + анализ
        # 3) Фамилия + анализ (с учётом склонения фамилии по полу)
        # 4) ФИО + М/О (М/О, М\О, МО)
        # Если нашли на более сильном этапе, дальше не ищем.
        match_entries = []
        match_strategy = "-"

        def collect_matches(predicate, strategy_label):
            local_entries = []
            local_seen_rows = set()

            for prepared in prepared_my_records:
                item = prepared["item"]

                if not predicate(prepared):
                    continue

                row_key = item.get("row")
                if row_key in local_seen_rows:
                    continue

                local_seen_rows.add(row_key)
                local_entries.append((item, strategy_label))

            return local_entries

        # 1) ФИО + анализ
        match_entries = collect_matches(
            lambda prepared: prepared["fio_norm"] == full_fio_key and analysis_keys_match(prepared["item"].get("analysis"), analysis),
            "1: ФИО+анализ",
        )
        if match_entries:
            match_strategy = "1: ФИО+анализ"
        else:
            # 2) Фамилия + Имя + анализ
            if details_surname_name_key:
                match_entries = collect_matches(
                    lambda prepared: prepared["surname_name_key"] == details_surname_name_key and analysis_keys_match(prepared["item"].get("analysis"), analysis),
                    "2: Фамилия+Имя+анализ",
                )
            if match_entries:
                match_strategy = "2: Фамилия+Имя+анализ"
            else:
                # 3) Фамилия + анализ (с учётом склонения фамилии по полу)
                if details_surname_gender_key:
                    match_entries = collect_matches(
                        lambda prepared: prepared["surname_gender_key"] == details_surname_gender_key and analysis_keys_match(prepared["item"].get("analysis"), analysis),
                        "3: Фамилия+анализ",
                    )
                if match_entries:
                    match_strategy = "3: Фамилия+анализ"
                else:
                    # 4) ФИО + М/О (М/О, М\О, МО)
                    # Ищем М/О и в анализе, и в сумме — в my_table оно может стоять в любой из этих колонок.
                    match_entries = collect_matches(
                        lambda prepared: prepared["fio_norm"] == full_fio_key and (
                            is_mo_marker(prepared["item"].get("analysis"))
                            or is_mo_marker(prepared["item"].get("sum"))
                        ),
                        "4: ФИО+М/О",
                    )
                    if match_entries:
                        match_strategy = "4: ФИО+М/О"

        # Начиная с 6 столбца: ФИО | дата | анализ | сумма
        out_col = 6
        for idx, (m, strategy_label) in enumerate(match_entries, start=1):
            match_fio = f"{cell_str(m.get('fio'))} ({strategy_label})"
            match_date = format_output_date(m.get("date"))
            match_analysis = cell_str(m.get("analysis"))
            match_sum = m.get("sum")

            # Жёстко записываем по колонкам:
            # 6=ФИО, 7=дата, 8=анализ, 9=сумма
            fio_cell = ws.cell(row=row_num, column=out_col)
            date_cell = ws.cell(row=row_num, column=out_col + 1)
            analysis_cell = ws.cell(row=row_num, column=out_col + 2)
            sum_cell = ws.cell(row=row_num, column=out_col + 3)

            fio_cell.value = match_fio
            date_cell.value = match_date
            analysis_cell.value = match_analysis
            sum_cell.value = match_sum

            # Ссылка на строку во втором листе
            target_row = m["row"]
            fio_cell.hyperlink = f"#'{copied_sheet_name}'!B{target_row}"
            fio_cell.font = link_font

            if (
                isinstance(google_sum, (int, float))
                and isinstance(match_sum, (int, float))
                and match_sum < google_sum
            ):
                sum_cell.fill = red_fill

            out_col += 4

        max_match_cols = max(max_match_cols, len(match_entries) * 4)
        row_num += 1

    # Заголовки для совпадений
    col = 6
    match_num = 1
    while col < 6 + max_match_cols:
        for title in [f"ФИО {match_num}", f"Дата {match_num}", f"Анализ {match_num}", f"Сумма {match_num}"]:
            c = ws.cell(row=1, column=col, value=title)
            c.fill = header_fill
            c.font = bold_font
            col += 1
        match_num += 1

    # Включаем автофильтр на всю строку заголовков.
    # Тогда в Excel фильтр уже будет доступен, в том числе для столбца F и далее.
    total_last_col = max(5, 5 + max_match_cols)
    ws.auto_filter.ref = f"A1:{get_column_letter(total_last_col)}1"

    # Ширины
    widths = {
        "A": 35,
        "B": 14,
        "C": 20,
        "D": 14,
        "E": 14,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    for col_idx in range(6, 6 + max_match_cols):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Копия листа из my_table вторым листом
    # Копируем только реально заполненные строки и столбцы.
    source_wb, temp_copy_path = open_my_table_workbook_for_copy(my_table_source)
    try:
        copied_ws = copy_sheet_to_workbook(source_wb, my_sheet_name, wb, copied_sheet_name)

        if copied_ws.max_row == 0:
            copied_ws.cell(row=1, column=1, value="Исходный лист оказался пустым после копирования")

        wb.save(result_file)
    finally:
        try:
            source_wb.close()
        except Exception:
            pass
        if temp_copy_path:
            try:
                os.remove(temp_copy_path)
            except OSError:
                pass

def main():
    actual_my_table_source = resolve_my_table_source(MY_TABLE_DOWNLOAD_URL)
    print(f"Читаю my_table: {actual_my_table_source}")
    print(f"Лист my_table: {MY_SHEET_NAME}")
    print(f"Диапазон строк my_table: start={MY_TABLE_START_ROW}, end={MY_TABLE_END_ROW}")
    try:
        my_records = parse_my_table(
            actual_my_table_source,
            MY_SHEET_NAME,
            start_row=MY_TABLE_START_ROW,
            end_row=MY_TABLE_END_ROW,
        )
    except RuntimeError as e:
        print(f"Ошибка чтения my_table: {e}")
        print("Подсказка: теперь код работает только через Microsoft Graph delegated OAuth. Установи msal, заполни AZURE_CLIENT_ID и войди в тот Microsoft-аккаунт, у которого есть доступ к my_table. Ошибка про reserved scopes исправляется удалением offline_access из scopes device code flow.")
        return

    print(f"Читаю details_table из файла: {DETAILS_TABLE_FILE}")
    print(f"Лист details_table: {DETAILS_TABLE_SHEET_NAME}")
    print(f"Диапазон строк details_table: start=авто (первая строка, где в 1-м столбце стоит 1), fallback={DETAILS_TABLE_START_ROW}, end={DETAILS_TABLE_END_ROW}")
    details_records = parse_details_table(
        DETAILS_TABLE_FILE,
        DETAILS_TABLE_SHEET_NAME,
        start_row=DETAILS_TABLE_START_ROW,
        end_row=DETAILS_TABLE_END_ROW,
    )
    print("Читаю основной прайс из price.xls, лист 1...")
    price_map = parse_price_table(PRICE_FILE, PRICE_MAIN_SHEET_NAME)
    print("Читаю дополнительный прайс из price.xls, лист 2...")
    price2_map = parse_price2_table(PRICE_FILE, PRICE_EXTRA_SHEET_NAME)

    print(f"Моя таблица: {len(my_records)} записей")
    print(f"details_table: {len(details_records)} записей")
    print(f"Прайс лист 1: {len(price_map)} анализов")
    print(f"Прайс лист 2: {len(price2_map)} анализов")

    print("Начинаю сбор результирующего Excel...")
    print("Копия второго листа будет сохранена только по реально заполненным строкам.")
    print(
        "Цена берётся из price.xls: сначала с листа 1, если не найдено — с листа 2. Для цены из прайса используется только строгое совпадение нормализованного анализа без мягкого поиска по цифрам: например, 505Б не должен подставлять цену от 505. Сопоставление с my_table для людей и анализов идёт по шагам: 1) ФИО+анализ, 2) Фамилия+Имя+анализ, 3) Фамилия+анализ, 4) ФИО+М/О. Для сопоставления с my_table дополнительно разрешены соответствия: PRS1-INV = PRS, ОБС103 = ОБС 103, 511ГИЭСПБ = ФГДС+ТЕСТ, 457К1ПАТ-А = 457, 377С-УРО = 377, 377УРО = 377, 518СПБ = 518, 301УРО = 301, 302УРО = 302, 305УРО = 305, 308УРО = 308, 3090УРО = 3090, 343УРО = 343, 441-А = 441, 446-А = 446, 459-А = 459, 459 НОС = 459, 459 РОТ = 459, 467-А = 467, 159ЯГ = 159, 321СВ = 321, 11HOMA = 11НОМА, 1601ОСТ = 1601, 6802PH = 6802. Перед сопоставлением исходный текст ФИО в my_table дополнительно чистится от типовых опечаток, например АЛЕ САНДР -> АЛЕКСАНДР и АЛЕКСАНДРОВИЧЯ -> АЛЕКСАНДРОВИЧ. В итоговом листе включается Excel-фильтр, а найденные совпадения записываются справа в порядке: ФИО, дата, анализ, сумма."
    )
    build_result_file(
        details_records=details_records,
        my_records=my_records,
        price_map=price_map,
        price2_map=price2_map,
        my_table_source=actual_my_table_source,
        my_sheet_name=MY_SHEET_NAME,
        result_file=RESULT_FILE,
    )

    print(
        f"Готово. Результат сохранён в файл: {RESULT_FILE}. Если файл уже был открыт в Excel, закрой его и открой заново, иначе можно видеть старую версию.")

    try:
        if os.name == "nt":
            os.startfile(RESULT_FILE)
        else:
            import subprocess
            subprocess.Popen(["open", RESULT_FILE])
    except Exception as e:
        print(f"Не удалось автоматически открыть итоговый файл: {e}")


def copy_sheet_to_workbook(source_wb, source_sheet_name, target_wb, target_sheet_name):
    """
    Копирует лист из исходной книги в результирующую книгу без обхода гигантского
    формально-занятого диапазона. Берём только реально заполненные значения.
    """
    source_ws = source_wb[source_sheet_name]
    target_ws = target_wb.create_sheet(title=target_sheet_name)

    copied_rows = 0
    max_used_col = 1

    for row_index, row in enumerate(source_ws.iter_rows(values_only=True), start=1):
        # Удаляем хвост из пустых ячеек справа
        row_values = list(row)
        last_non_empty = 0
        for i in range(len(row_values) - 1, -1, -1):
            value = row_values[i]
            if value is not None and str(value).strip() != "":
                last_non_empty = i + 1
                break

        if last_non_empty == 0:
            # Полностью пустые строки не копируем
            continue

        trimmed_row = row_values[:last_non_empty]
        for col_index, value in enumerate(trimmed_row, start=1):
            target_ws.cell(row=row_index, column=col_index, value=value)

        copied_rows += 1
        max_used_col = max(max_used_col, last_non_empty)


    # Базовая автоподстройка ширины только по реально использованным столбцам
    for col_idx in range(1, max_used_col + 1):
        target_ws.column_dimensions[get_column_letter(col_idx)].width = 18

    print(f"Копирование листа завершено. Скопировано заполненных строк: {copied_rows}")
    return target_ws
# 2) создай app registration в Azure / Entra
# 3) укажи AZURE_CLIENT_ID
# 4) оставь в MY_TABLE_DOWNLOAD_URL обычную OneDrive share-ссылку
# 5) при первом запуске войди в Microsoft-аккаунт, у которого есть доступ к файлу
# details_table подгружается локально из DETAILS_TABLE_FILE.

if __name__ == "__main__":
    open_analyschecker_standalone()
    def get_conclusion_data_path(self):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_dir, "conclusion_form", "res", "data.xml")

    def load_conclusion_records(self):
        xml_path = self.get_conclusion_data_path()
        if not os.path.exists(xml_path):
            raise FileNotFoundError(f"Файл данных не найден: {xml_path}")

        tree = ET.parse(xml_path)
        root = tree.getroot()
        records = []

        for person in root.findall("person"):
            record = {
                "organization": person.findtext("organization", default=""),
                "name": person.findtext("name", default=""),
                "birthday": person.findtext("birthday", default=""),
                "sex": person.findtext("sex", default=""),
                "division": person.findtext("division", default=""),
                "profession": person.findtext("profession", default=""),
                "factors": person.findtext("factors", default=""),
                "typework": person.findtext("typework", default=""),
                "ids_date": person.findtext("ids_date", default=""),
                "diagnosis": person.findtext("diagnosis", default=""),
                "card_number": person.findtext("card_number", default=""),
                "id": person.findtext("id", default=""),
                "date_created": person.findtext("date_created", default=""),
            }
            records.append(record)

        return records

    def open_conclusions_editor(self):
        try:
            all_records = self.load_conclusion_records()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить заключения:\n{e}")
            return

        top = tk.Toplevel(self)
        top.title("Редактировать заключения")
        top.geometry("1500x700")

        filter_frame = tk.Frame(top)
        filter_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(filter_frame, text="Период с:").grid(row=0, column=0, sticky="w")
        start_var = tk.StringVar()
        start_entry = tk.Entry(filter_frame, textvariable=start_var, width=16)
        start_entry.grid(row=0, column=1, padx=(6, 12), sticky="w")

        tk.Label(filter_frame, text="По:").grid(row=0, column=2, sticky="w")
        end_var = tk.StringVar()
        end_entry = tk.Entry(filter_frame, textvariable=end_var, width=16)
        end_entry.grid(row=0, column=3, padx=(6, 12), sticky="w")

        status_var = tk.StringVar(value=f"Загружено записей: {len(all_records)}")
        tk.Label(filter_frame, textvariable=status_var, fg="#555").grid(row=0, column=6, padx=(12, 0), sticky="w")

        table_frame = tk.Frame(top)
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        columns = (
            "organization", "name", "birthday", "sex", "division", "profession",
            "factors", "typework", "ids_date", "diagnosis", "card_number", "id", "date_created"
        )
        headings = {
            "organization": "Организация",
            "name": "ФИО",
            "birthday": "Дата рождения",
            "sex": "Пол",
            "division": "Подразделение",
            "profession": "Должность",
            "factors": "Факторы",
            "typework": "Виды работ",
            "ids_date": "Дата ИДС",
            "diagnosis": "Диагноз",
            "card_number": "№ карты",
            "id": "ID",
            "date_created": "Создано",
        }

        tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        width_map = {
            "organization": 180,
            "name": 220,
            "birthday": 110,
            "sex": 60,
            "division": 150,
            "profession": 150,
            "factors": 220,
            "typework": 220,
            "ids_date": 100,
            "diagnosis": 180,
            "card_number": 110,
            "id": 100,
            "date_created": 140,
        }

        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=width_map.get(col, 120), anchor="w")

        def parse_ddmmyyyy(value):
            text = cell_str(value)
            if not text:
                return None
            try:
                return datetime.strptime(text, "%d.%m.%Y")
            except ValueError:
                return None

        def fill_tree(records):
            for item in tree.get_children():
                tree.delete(item)

            for rec in records:
                tree.insert(
                    "",
                    "end",
                    values=tuple(rec.get(col, "") for col in columns)
                )

            status_var.set(f"Показано записей: {len(records)}")

        def apply_period_filter():
            start_text = start_var.get().strip()
            end_text = end_var.get().strip()

            start_dt = parse_ddmmyyyy(start_text) if start_text else None
            end_dt = parse_ddmmyyyy(end_text) if end_text else None

            if start_text and start_dt is None:
                messagebox.showerror("Ошибка", "Дата 'Период с' должна быть в формате ДД.ММ.ГГГГ")
                return
            if end_text and end_dt is None:
                messagebox.showerror("Ошибка", "Дата 'По' должна быть в формате ДД.ММ.ГГГГ")
                return
            if start_dt and end_dt and end_dt < start_dt:
                messagebox.showerror("Ошибка", "Дата конца периода не может быть меньше даты начала")
                return

            filtered = []
            for rec in all_records:
                ids_dt = parse_ddmmyyyy(rec.get("ids_date", ""))
                if start_dt and (ids_dt is None or ids_dt < start_dt):
                    continue
                if end_dt and (ids_dt is None or ids_dt > end_dt):
                    continue
                filtered.append(rec)

            fill_tree(filtered)

        def reset_period_filter():
            start_var.set("")
            end_var.set("")
            fill_tree(all_records)

        tk.Button(filter_frame, text="Показать", command=apply_period_filter).grid(row=0, column=4, padx=(0, 8))
        tk.Button(filter_frame, text="Сбросить", command=reset_period_filter).grid(row=0, column=5, padx=(0, 8))

        fill_tree(all_records)