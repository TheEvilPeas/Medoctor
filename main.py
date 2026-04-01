import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkinter.ttk import Combobox
from openpyxl import load_workbook
import os
import json
import datetime
import calendar
import pandas as pd
from conclusion_form.form import ConclusionForm
from search_form.form import SearchForm
from analyschecker_form.form import AnalysCheckerForm
import sys, os, json


def setup_logging():
    log_dir = os.getcwd()
    log_file = os.path.join(log_dir, "log.txt")

    # Чтобы старые логи не затирались, можно дописывать
    sys.stdout = open(log_file, "a", encoding="utf-8")
    sys.stderr = sys.stdout

    print("\n=== Запуск программы:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "===\n")
    print("cwd:", os.getcwd(), flush=True)
    print("sys.executable:", sys.executable, flush=True)
    print("__file__:", __file__, flush=True)


# --- Debug log helper ---
def debug_log(*parts):
    try:
        print("[DEBUG]", *parts, flush=True)
    except Exception:
        pass

def settings_path():
    path = resource_path("res/settings.json")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    return path

def resource_path(rel_path: str) -> str:
    """
    Возвращает путь к ресурсу и в dev-режиме, и внутри PyInstaller.
    rel_path: относительный путь внутри проекта (например 'conclusion_form/res/template.docx')
    """
    if hasattr(sys, '_MEIPASS'):
        base = sys._MEIPASS  # временная папка PyInstaller
    else:
        base = os.path.abspath(".")
    return os.path.join(base, rel_path)

SETTINGS_PATH = settings_path()
XML_PATH      = resource_path("conclusion_form/res/data.xml")
PRIKAZ_XLSX   = resource_path("search_form/input/prikaz29n.xlsx")




def load_settings():
    if os.path.exists(SETTINGS_PATH):
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    else:
        return {"save_dir": os.getcwd()}

def save_settings(settings):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=2, ensure_ascii=False)

def load_data():
    import xml.etree.ElementTree as ET
    if not os.path.exists(XML_PATH):
        return {}
    tree = ET.parse(XML_PATH)
    root = tree.getroot()
    data = {}
    for person in root.findall("person"):
        org_name = person.findtext("organization", default="")
        record = {
            "type": person.findtext("type", default="предварительный"),
            "name": person.findtext("name", default=""),
            "birthday": person.findtext("birthday", default=""),
            "sex": person.findtext("sex", default=""),
            "division": person.findtext("division", default=""),
            "profession": person.findtext("profession", default=""),
            "factors": person.findtext("factors", default=""),
            "typework": person.findtext("typework", default=""),
            "id": person.findtext("id", default=""),
            "ids_date": person.findtext("ids_date", default=""),
            "diagnosis": person.findtext("diagnosis", default=""),
            "card_number": person.findtext("card_number", default="")
        }
        if org_name not in data:
            data[org_name] = []
        data[org_name].append(record)
    return data

def sanitize_filename(name: str) -> str:
    import re
    return re.sub(r'[\\\/\:\*\?"<>\|]', '_', name)

def is_valid_date(date_str):
    try:
        datetime.datetime.strptime(date_str, "%d.%m.%Y")
        return True
    except ValueError:
        return False

def open_calendar(parent, entry_widget):
    from tkcalendar import Calendar
    import ctypes
    from ctypes import wintypes

    mouse_x = parent.winfo_pointerx()
    mouse_y = parent.winfo_pointery()
    top = tk.Toplevel(parent)
    top.overrideredirect(False)
    top.title("Выберите дату")
    def pick_date():
        date = cal.selection_get()
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, date.strftime("%d.%m.%Y"))
        top.destroy()
    cal = Calendar(top, date_pattern='dd.mm.yyyy')
    cal.pack(padx=10, pady=10)
    tk.Button(top, text="Выбрать", command=pick_date).pack(pady=5)
    top.update_idletasks()
    win_w = top.winfo_width()
    win_h = top.winfo_height()
    try:
        SPI_GETWORKAREA = 0x0030
        rect = wintypes.RECT()
        ctypes.windll.user32.SystemParametersInfoW(SPI_GETWORKAREA, 0, ctypes.byref(rect), 0)
        work_w = rect.right - rect.left
        work_h = rect.bottom - rect.top
    except Exception:
        work_w = parent.winfo_screenwidth()
        work_h = parent.winfo_screenheight()
    x = mouse_x + 10
    y = mouse_y + 10
    if x + win_w > work_w:
        x = work_w - win_w
    if y + win_h > work_h:
        y = work_h - win_h
    if x < 0:
        x = 0
    if y < 0:
        y = 0
    top.geometry(f"{win_w}x{win_h}+{x}+{y}")

def show_notification(parent, text, duration=3000, x_offset=10, y_offset=10):
    label = tk.Label(
        parent,
        text=text,
        fg="white",
        bg="#333",
        bd=1,
        relief="solid",
        padx=10, pady=5
    )
    label.place(
        relx=1.0, rely=1.0,
        anchor="se",
        x=-x_offset,
        y=-y_offset
    )
    parent.after(duration, label.place_forget)

class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Генератор медицинских заключений")
        self.geometry("680x750")
        self.apply_light_theme()
        self.settings = load_settings()
        self.current_form_frame = None

        self.create_menubar()
        self.settings_window = None
        self.prikaz_replace_window = None
        self.forms_area = tk.Frame(self)
        self.forms_area.pack(fill="both", expand=True)
        self.show_form("search")




    def apply_light_theme(self):
        self.configure(bg="white")
        self.tk_setPalette(
            background="white",
            foreground="black",
            activeBackground="#e9e9e9",
            activeForeground="black"
        )

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("TFrame", background="white")
        style.configure("TLabel", background="white", foreground="black")
        style.configure("TButton", background="white", foreground="black")
        style.configure("TEntry", fieldbackground="white", foreground="black")
        style.configure("TCombobox", fieldbackground="white", background="white", foreground="black")

    def create_menubar(self):
        menubar = tk.Menu(self)

        # Файл (настройки)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Настройки", command=self.open_settings)
        menubar.add_cascade(label="Файл", menu=file_menu)

        # Формы (переключение форм через меню)
        forms_menu = tk.Menu(menubar, tearoff=0)
        forms_menu.add_command(label="Заключение", command=lambda: self.show_form("conclusion"))
        forms_menu.add_command(label="Анализ по приказу 29н", command=lambda: self.show_form("search"))
        forms_menu.add_command(label="Проверка анализов", command=lambda: self.show_form("analyschecker"))
        menubar.add_cascade(label="Формы", menu=forms_menu)

        # Отчеты
        reports_menu = tk.Menu(menubar, tearoff=0)
        reports_menu.add_command(label="Отчёт по организации", command=self.report_by_organization)
        reports_menu.add_command(label="Отчёт за месяц", command=self.report_by_month)
        reports_menu.add_command(label="Отчёт по врачам", command=self.report_doctors)
        reports_menu.add_command(label="Кастомный отчёт", command=self.report_custom)
        menubar.add_cascade(label="Отчёты", menu=reports_menu)

        self.config(menu=menubar)

    def report_doctors(self):
        import re, os, datetime, calendar
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from tkinter import messagebox

        data = load_data()

        # --- Окно выбора дат (как в отчёте за месяц) ---
        rpt = tk.Toplevel(self)
        rpt.title("Отчет по врачам")
        rpt.resizable(False, False)
        padx, pady = 10, 5

        tk.Label(rpt, text="Период с:").grid(row=0, column=0, sticky="w", padx=padx, pady=pady)
        start_var = tk.StringVar()
        start_entry = tk.Entry(rpt, width=20, textvariable=start_var)
        start_entry.grid(row=0, column=1, sticky="w", padx=padx, pady=pady)
        tk.Button(rpt, text="📅", command=lambda: open_calendar(self, start_entry)).grid(row=0, column=2, padx=0,
                                                                                        pady=pady)

        tk.Label(rpt, text="По:").grid(row=1, column=0, sticky="w", padx=padx, pady=pady)
        end_var = tk.StringVar()
        end_entry = tk.Entry(rpt, width=20, textvariable=end_var)
        end_entry.grid(row=1, column=1, sticky="w", padx=padx, pady=pady)
        tk.Button(rpt, text="📅", command=lambda: open_calendar(self, end_entry)).grid(row=1, column=2, padx=0,
                                                                                      pady=pady)

        def on_start_changed(*_):
            s = start_var.get().strip()
            if is_valid_date(s):
                dt = datetime.datetime.strptime(s, "%d.%m.%Y")
                last = calendar.monthrange(dt.year, dt.month)[1]
                end_var.set(f"{last:02d}.{dt.month:02d}.{dt.year}")

        start_var.trace_add("write", on_start_changed)

        # -------- helpers --------
        def parse_points_from_text(text: str):
            if not text:
                return []
            norm = text.replace(';', ',')
            return [p.strip() for p in re.findall(r'\d+(?:\.\d+)?', norm) if p.strip()]

        def calc_age_on(date_birth: str, at_date: str):
            try:
                b = datetime.datetime.strptime(date_birth, "%d.%m.%Y").date()
                d = datetime.datetime.strptime(at_date, "%d.%m.%Y").date()
            except Exception:
                return None
            return d.year - b.year - ((d.month, d.day) < (b.month, b.day))

        def base_point_for_gender_age(sex: str, age):
            if age is None:
                return None
            if (sex or "").strip().upper() == "М":
                return "0.12" if age >= 40 else "0.11"
            else:
                return "0.22" if age >= 40 else "0.21"

        def make_report_doctors():
            start = start_var.get().strip()
            end = end_var.get().strip()

            if not is_valid_date(start) or not is_valid_date(end):
                messagebox.showerror("Ошибка даты", "Даты в формате ДД.ММ.ГГГГ")
                return
            d0 = datetime.datetime.strptime(start, "%d.%m.%Y")
            d1 = datetime.datetime.strptime(end, "%d.%m.%Y")
            if d1 < d0:
                messagebox.showerror("Ошибка", "Конечная дата меньше начальной")
                return

            map_path = PRIKAZ_XLSX
            if not os.path.exists(map_path):
                messagebox.showerror("Ошибка", f"Не найден файл: {map_path}")
                return
            df_map = pd.read_excel(map_path)
            df_map['n'] = df_map['n'].astype(str).str.replace(',', '.').str.strip()
            def split_to_values(series):
                items = []
                for v in series.fillna("").astype(str):
                    raw_text = str(v).replace("\n", ",").replace(";", ",")
                    parts = [p.strip() for p in raw_text.split(",") if p.strip()]
                    items.extend(parts)
                return items

            def unique_preserve_order(values):
                result = []
                seen = set()
                for value in values:
                    key = value.lower()
                    if key in seen:
                        continue
                    seen.add(key)
                    result.append(value)
                return result

            doctor_cols = unique_preserve_order(split_to_values(df_map.get('doctors_name', pd.Series(dtype=str))))
            target_cols = doctor_cols + [
                "ФОГ",
                "Мамография",
                "Спирометрия",
                "Тональная пороговая аудиометрия",
            ]

            test_patterns = {
                "ФОГ": ["фог", "флюорограф", "флюорография", "рентген грудной"],
                "Мамография": ["маммограф", "маммография"],
                "Спирометрия": ["спирометр"],
                "Тональная пороговая аудиометрия": ["аудиометр", "тональная пороговая аудиометрия"],
            }

            def split_to_set(series):
                return set(v.lower() for v in split_to_values(series) if v)

            def contains_any(terms_set, patterns):
                return any(p in t for t in terms_set for p in patterns)

            summary_rows = []
            for org_name, recs in data.items():
                for r in recs:
                    ids = (r.get("ids_date") or "").strip()
                    if not ids:
                        continue
                    try:
                        d_ids = datetime.datetime.strptime(ids, "%d.%m.%Y")
                    except ValueError:
                        continue
                    if not (d0 <= d_ids <= d1):
                        continue

                    age = calc_age_on(r.get("birthday", ""), ids)
                    base_pt = base_point_for_gender_age(r.get("sex", ""), age)

                    pts = []
                    pts += parse_points_from_text(r.get("factors", ""))
                    pts += parse_points_from_text(r.get("typework", ""))
                    if base_pt:
                        pts.append(base_pt)

                    items = sorted(set(p for p in pts if p))
                    subset = df_map[df_map['n'].isin(items)] if items else df_map.iloc[0:0]

                    required_doctors = split_to_set(subset.get('doctors_name', pd.Series(dtype=str)))
                    required_inspections = split_to_set(subset.get('inspection', pd.Series(dtype=str)))
                    required_analyses = split_to_set(subset.get('analysis', pd.Series(dtype=str)))

                    row = {
                        "Дата": ids,
                        "Тип осмотра": r.get("type", "предварительный"),
                        "ФИО": r.get("name", ""),
                        "Дата рождения": r.get("birthday", ""),
                        "Организация": org_name,
                    }
                    for col in doctor_cols:
                        row[col] = '+' if col.lower() in required_doctors else ''
                    for col, pats in test_patterns.items():
                        has = contains_any(required_inspections, pats) or contains_any(required_analyses, pats)
                        row[col] = '+' if has else ''
                    summary_rows.append(row)

            if not summary_rows:
                messagebox.showinfo("Пустой отчет", "Нет записей за выбранный период.")
                return

            summary_rows.sort(key=lambda x: datetime.datetime.strptime(x["Дата"], "%d.%m.%Y"))
            columns = ["Дата", "Тип осмотра", "ФИО", "Дата рождения", "Организация"] + target_cols
            df = pd.DataFrame(summary_rows, columns=columns)

            save_dir = self.settings.get("save_dir", os.getcwd())
            save_path = os.path.join(save_dir, sanitize_filename(f"Отчет_по_врачам_{start}_{end}.xlsx"))

            try:
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment
                from openpyxl import load_workbook

                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Отчет")

                wb = load_workbook(save_path)
                ws = wb["Отчет"]

                # Высота первой строки
                ws.row_dimensions[1].height = 136

                for col_idx, col_name in enumerate(columns, 1):
                    # Вертикальный текст для всех колонок
                    ws.cell(row=1, column=col_idx).alignment = Alignment(
                        textRotation=90,
                        vertical="center",
                        horizontal="center",
                        wrap_text=True
                    )
                    if col_name in target_cols:
                        # Мед. колонки фикс ширина 60
                        ws.column_dimensions[get_column_letter(col_idx)].width = 8
                    else:
                        # Остальные — автоподбор
                        max_len = max(
                            (len(str(cell.value)) if cell.value else 0) for cell in ws[get_column_letter(col_idx)])
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

                wb.save(save_path)
                show_notification(self, f"Отчет сохранён: {save_path}")
                rpt.destroy()

            except Exception as e:
                messagebox.showerror("Ошибка записи", f"Не удалось сохранить файл:\n{e}")

        tk.Button(
            rpt,
            text="Сформировать",
            command=make_report_doctors,
            bg="#4CAF50",
            fg="white"
        ).grid(row=2, column=0, columnspan=3, pady=(10, 10), padx=padx, sticky="ew")

        rpt.grid_columnconfigure(1, weight=1)
        rpt.update_idletasks()
        w, h = rpt.winfo_width(), rpt.winfo_height()
        sw, sh = rpt.winfo_screenwidth(), rpt.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 2
        rpt.geometry(f"{w}x{h}+{x}+{y}")

    def create_forms_panel(self):
        panel = tk.Frame(self, bd=1, relief="raised")
        panel.pack(fill="x")
        tk.Label(panel, text="Формы:", font="Arial 10 bold").pack(side="left", padx=10, pady=5)
        tk.Button(panel, text="Заключение", command=lambda: self.show_form("conclusion")).pack(side="left", padx=5, pady=5)

    def show_form(self, form_key):
        if self.current_form_frame is not None:
            try:
                self.current_form_frame.destroy()
            except Exception:
                pass
            self.current_form_frame = None

        for child in self.forms_area.winfo_children():
            try:
                child.destroy()
            except Exception:
                pass

        if form_key == "conclusion":
            self.current_form_frame = ConclusionForm(self.forms_area, main_app=self)
            self.current_form_frame.pack(fill="both", expand=True)
        elif form_key == "search":
            self.current_form_frame = SearchForm(self.forms_area, main_app=self)
            self.current_form_frame.pack(fill="both", expand=True)
        elif form_key == "analyschecker":
            self.current_form_frame = AnalysCheckerForm(self.forms_area, main_app=self)
            self.current_form_frame.pack(fill="both", expand=True)

    def open_settings(self):
        # если окно уже существует — поднять и сфокусировать
        if self.settings_window and self.settings_window.winfo_exists():
            self.settings_window.deiconify()
            self.settings_window.lift()
            self.settings_window.focus_force()
            # кратко сделать topmost, чтобы наверняка всплыло
            self.settings_window.attributes('-topmost', True)
            self.settings_window.after(100, lambda: self.settings_window.attributes('-topmost', False))
            return

        # иначе — создать новое и сохранить ссылку
        top = tk.Toplevel(self)
        self.settings_window = top
        top.title("Настройки")
        top.resizable(False, False)
        top.geometry("760x220")

        # при закрытии окна обнуляем ссылку
        def on_close():
            if self.settings_window and self.settings_window.winfo_exists():
                self.settings_window.destroy()
            self.settings_window = None

        top.protocol("WM_DELETE_WINDOW", on_close)

        tk.Label(top, text="Папка для сохранения документов:").pack(anchor="w", padx=10, pady=(10, 0))
        path_var = tk.StringVar(value=self.settings.get("save_dir", os.getcwd()))

        path_row = tk.Frame(top)
        path_row.pack(fill="x", padx=10, pady=5)
        path_row.grid_columnconfigure(0, weight=1)

        path_entry = tk.Entry(path_row, textvariable=path_var, width=50)
        path_entry.grid(row=0, column=0, sticky="ew")

        def select_directory():
            from tkinter import filedialog
            # временно убираем topmost, чтобы диалог не оказался за окном
            self.settings_window.attributes("-topmost", False)
            path = filedialog.askdirectory(parent=self.settings_window)
            # после выбора — вернуть окно наверх
            self.settings_window.lift()
            self.settings_window.focus_force()
            self.settings_window.attributes("-topmost", True)
            self.settings_window.after(100, lambda: self.settings_window.attributes("-topmost", False))

            if path:
                path_var.set(path)

        def save_and_close():
            selected_path = path_var.get()
            if not os.path.exists(selected_path):
                try:
                    os.makedirs(selected_path)
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Не удалось создать папку:\n{e}")
                    return
            self.settings["save_dir"] = selected_path
            save_settings(self.settings)
            on_close()

        tk.Button(path_row, text="Выбрать...", command=select_directory).grid(row=0, column=1, padx=(8, 0))
        tk.Button(path_row, text="Сохранить", command=save_and_close).grid(row=0, column=2, padx=(8, 0))

        # tk.Button(top, text="Редактировать приказ 29н…", command=self.open_prikaz_for_edit).pack(pady=(5, 0))
        tk.Button(top, text="Изменить названия в приказе 29н…", command=self.open_prikaz_replace_window).pack(pady=(15, 10))

    def get_prikaz_first_sheet_and_headers(self):
        debug_log("get_prikaz_first_sheet_and_headers() called")
        debug_log("PRIKAZ_XLSX =", PRIKAZ_XLSX)
        debug_log("exists =", os.path.exists(PRIKAZ_XLSX))

        if not os.path.exists(PRIKAZ_XLSX):
            raise FileNotFoundError(f"Не найден файл: {PRIKAZ_XLSX}")

        wb = load_workbook(PRIKAZ_XLSX)
        ws = wb[wb.sheetnames[0]]
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header is None:
                continue
            header = str(header).strip()
            if header:
                headers[header] = col_idx

        debug_log("sheetnames =", wb.sheetnames)
        debug_log("active first sheet =", ws.title)
        debug_log("headers =", list(headers.keys()))
        debug_log("max_row =", ws.max_row, "max_column =", ws.max_column)
        return wb, ws, headers

    def get_prikaz_replace_columns(self):
        _, _, headers = self.get_prikaz_first_sheet_and_headers()
        preferred = ["doctors_name", "inspection", "analysis"]
        return [name for name in preferred if name in headers]

    def get_unique_prikaz_values(self, column_name):
        debug_log("get_unique_prikaz_values() called with column_name =", column_name)
        wb, ws, headers = self.get_prikaz_first_sheet_and_headers()
        try:
            if column_name not in headers:
                debug_log("column not found in headers:", column_name)
                return []

            col_idx = headers[column_name]
            debug_log("column index =", col_idx)
            unique_values = set()
            sample_raw = []

            for row_idx in range(2, ws.max_row + 1):
                raw = ws.cell(row=row_idx, column=col_idx).value
                if len(sample_raw) < 10:
                    sample_raw.append((row_idx, raw))
                if raw is None:
                    continue

                raw_text = str(raw).replace("\n", ",").replace(";", ",")
                for part in raw_text.split(','):
                    value = " ".join(part.strip().split())
                    if value:
                        unique_values.add(value)

            debug_log("sample raw values =", sample_raw)
            result = sorted(unique_values, key=lambda x: x.lower())
            debug_log("unique values count =", len(result))
            debug_log("first unique values =", result[:20])
            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def replace_prikaz_value(self, column_name, old_value, new_value):
        import tempfile
        import os

        debug_log("replace_prikaz_value() called")
        debug_log("column_name =", column_name)
        debug_log("old_value =", repr(old_value))
        debug_log("new_value =", repr(new_value))
        debug_log("PRIKAZ_XLSX =", PRIKAZ_XLSX)

        wb, ws, headers = self.get_prikaz_first_sheet_and_headers()
        if column_name not in headers:
            wb.close()
            raise ValueError(f"Колонка '{column_name}' не найдена в первом листе приказа")

        def normalize_value(text):
            return " ".join(str(text).replace("\n", " ").replace(";", " ").strip().split()).lower()

        col_idx = headers[column_name]
        changed = 0
        summer_changed = 0
        old_norm = normalize_value(old_value)
        new_clean = " ".join(new_value.strip().split())
        debug_log("normalized old_value =", repr(old_norm))
        debug_log("normalized new_value =", repr(new_clean))

        matches_preview = []

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            raw = cell.value
            if raw is None:
                continue

            original_text = str(raw)
            original_norm = normalize_value(original_text)

            if len(matches_preview) < 20:
                matches_preview.append((row_idx, original_text, original_norm))

            # 1) Если вся ячейка целиком совпадает со старым значением — меняем целиком
            if original_norm == old_norm:
                debug_log("full cell match at row", row_idx, "old=", repr(original_text), "new=", repr(new_clean))
                cell.value = new_clean
                changed += 1
                continue

            # 2) Иначе пробуем заменить внутри списка значений, разделённых запятыми / ; / переносами
            raw_text = original_text.replace("\n", ",").replace(";", ",")
            parts = [part.strip() for part in raw_text.split(",") if part.strip()]
            replaced_any = False
            new_parts = []

            for part in parts:
                part_norm = normalize_value(part)
                if part_norm == old_norm:
                    debug_log("partial match at row", row_idx, "part=", repr(part), "new=", repr(new_clean))
                    new_parts.append(new_clean)
                    replaced_any = True
                else:
                    new_parts.append(" ".join(part.split()))

            if replaced_any:
                new_cell_value = ", ".join(new_parts)
                debug_log("row", row_idx, "updated from", repr(original_text), "to", repr(new_cell_value))
                cell.value = new_cell_value
                changed += 1

        debug_log("matches preview =", matches_preview)
        debug_log("changed rows total on first sheet =", changed)

        # Дополнительно обновляем лист summer (2-й лист), колонка A
        if len(wb.sheetnames) >= 2:
            summer_ws = wb[wb.sheetnames[1]]
            debug_log("updating summer sheet =", summer_ws.title)
            for row_idx in range(2, summer_ws.max_row + 1):
                summer_cell = summer_ws.cell(row=row_idx, column=1)
                summer_raw = summer_cell.value
                if summer_raw is None:
                    continue

                summer_text = str(summer_raw)
                summer_norm = normalize_value(summer_text)
                if summer_norm == old_norm:
                    debug_log("summer match at row", row_idx, "old=", repr(summer_text), "new=", repr(new_clean))
                    summer_cell.value = new_clean
                    summer_changed += 1

        debug_log("changed rows total in summer =", summer_changed)

        total_changed = changed + summer_changed

        if total_changed > 0:
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            debug_log("temporary save path =", tmp_path)
            try:
                wb.save(tmp_path)
                debug_log("temporary file saved")
                wb.close()
                os.replace(tmp_path, PRIKAZ_XLSX)
                debug_log("original file replaced successfully")
            except Exception as e:
                debug_log("save/replace error =", repr(e))
                try:
                    wb.close()
                except Exception:
                    pass
                if os.path.exists(tmp_path):
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                raise
        else:
            debug_log("no changes found in first sheet and summer; workbook closed without save")
            wb.close()

        return total_changed

    def get_summer_sheet_rows(self):
        debug_log("get_summer_sheet_rows() called")
        if not os.path.exists(PRIKAZ_XLSX):
            raise FileNotFoundError(f"Не найден файл: {PRIKAZ_XLSX}")

        wb = load_workbook(PRIKAZ_XLSX, data_only=False)
        try:
            if len(wb.sheetnames) < 2:
                raise ValueError("Во втором листе файла приказа не найден лист summer")

            ws = wb[wb.sheetnames[1]]
            debug_log("summer sheet title =", ws.title)
            rows = []
            for row_idx in range(2, ws.max_row + 1):
                name = ws.cell(row=row_idx, column=1).value
                price = ws.cell(row=row_idx, column=2).value
                if name is None or str(name).strip() == "":
                    continue
                rows.append({
                    "row_idx": row_idx,
                    "name": str(name).strip(),
                    "price": "" if price is None else str(price).strip(),
                })
            debug_log("summer rows loaded =", len(rows))
            return rows
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def update_summer_price(self, row_idx, new_price):
        import tempfile
        import os

        debug_log("update_summer_price() called", "row_idx=", row_idx, "new_price=", repr(new_price))
        if not os.path.exists(PRIKAZ_XLSX):
            raise FileNotFoundError(f"Не найден файл: {PRIKAZ_XLSX}")

        wb = load_workbook(PRIKAZ_XLSX)
        try:
            if len(wb.sheetnames) < 2:
                raise ValueError("Во втором листе файла приказа не найден лист summer")

            ws = wb[wb.sheetnames[1]]
            if row_idx < 2 or row_idx > ws.max_row:
                raise ValueError(f"Некорректный номер строки листа summer: {row_idx}")

            price_text = str(new_price).strip().replace(",", ".")
            if price_text == "":
                raise ValueError("Цена не может быть пустой")

            price_value = float(price_text)
            if price_value.is_integer():
                price_value = int(price_value)

            ws.cell(row=row_idx, column=2).value = price_value

            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            try:
                wb.save(tmp_path)
                wb.close()
                os.replace(tmp_path, PRIKAZ_XLSX)
                debug_log("summer price updated successfully")
            except Exception:
                if os.path.exists(tmp_path):
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                raise
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def open_prikaz_replace_window(self):
        debug_log("open_prikaz_replace_window() called")

        if self.prikaz_replace_window and self.prikaz_replace_window.winfo_exists():
            debug_log("replace window already exists; focusing it")
            self.prikaz_replace_window.deiconify()
            self.prikaz_replace_window.lift()
            self.prikaz_replace_window.focus_force()
            return

        try:
            columns = self.get_prikaz_replace_columns()
            summer_rows = self.get_summer_sheet_rows()
            debug_log("replace columns =", columns)
        except Exception as e:
            debug_log("open_prikaz_replace_window error while reading data =", repr(e))
            messagebox.showerror("Ошибка", f"Не удалось прочитать приказ 29н:\n{e}")
            return

        if not columns:
            messagebox.showerror(
                "Ошибка",
                "В первом листе приказа не найдены колонки doctors_name, inspection или analysis."
            )
            return

        top = tk.Toplevel(self)
        self.prikaz_replace_window = top
        top.title("Переименование значений и цены summer")
        top.geometry("1180x620")
        top.minsize(980, 540)
        top.resizable(True, True)

        def on_close():
            debug_log("replace window closed")
            if self.prikaz_replace_window and self.prikaz_replace_window.winfo_exists():
                self.prikaz_replace_window.destroy()
            self.prikaz_replace_window = None

        top.protocol("WM_DELETE_WINDOW", on_close)

        container = tk.PanedWindow(top, orient=tk.HORIZONTAL, sashrelief="raised")
        container.pack(fill="both", expand=True, padx=10, pady=10)

        left_frame = tk.Frame(container)
        right_frame = tk.Frame(container)
        container.add(left_frame, minsize=420)
        container.add(right_frame, minsize=420)

        # -------- Левая часть: переименование значений --------
        tk.Label(left_frame, text="Колонка приказа:").pack(anchor="w")
        column_var = tk.StringVar(value=columns[0])
        column_cb = Combobox(left_frame, textvariable=column_var, values=columns, state="readonly", width=35)
        column_cb.pack(anchor="w", pady=(0, 10), fill="x")

        tk.Label(left_frame, text="Уникальные значения:").pack(anchor="w")
        list_frame = tk.Frame(left_frame)
        list_frame.pack(fill="both", expand=True)

        value_list = tk.Listbox(list_frame, exportselection=False, activestyle="dotbox")
        value_list.pack(side="left", fill="both", expand=True)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=value_list.yview)
        scrollbar.pack(side="right", fill="y")
        value_list.config(yscrollcommand=scrollbar.set)

        old_var = tk.StringVar()
        tk.Label(left_frame, text="Новое значение для выбранного элемента:").pack(anchor="w", pady=(10, 0))
        new_var = tk.StringVar()
        new_entry = tk.Entry(left_frame, textvariable=new_var)
        new_entry.pack(fill="x", pady=(0, 10))

        left_status_var = tk.StringVar(value=f"Файл: {PRIKAZ_XLSX}")
        tk.Label(left_frame, textvariable=left_status_var, fg="#555", anchor="w", justify="left").pack(fill="x", pady=(0, 8))

        left_button_frame = tk.Frame(left_frame)
        left_button_frame.pack(fill="x", pady=(4, 0))

        def refresh_values(*_):
            debug_log("refresh_values() called")
            value_list.delete(0, tk.END)
            old_var.set("")
            new_var.set("")

            column_name = column_var.get().strip()
            debug_log("selected column =", column_name)
            if not column_name:
                left_status_var.set("Не выбрана колонка")
                messagebox.showwarning("Внимание", "Сначала выбери колонку")
                return

            try:
                values = self.get_unique_prikaz_values(column_name)
            except Exception as e:
                debug_log("refresh_values error =", repr(e))
                left_status_var.set(f"Ошибка загрузки: {e}")
                messagebox.showerror("Ошибка", f"Не удалось загрузить значения:\n{e}\n\nФайл: {PRIKAZ_XLSX}")
                return

            if not values:
                debug_log("no values found for column", column_name)
                left_status_var.set(f"Колонка '{column_name}': значения не найдены")
                messagebox.showinfo(
                    "Значения не найдены",
                    f"В колонке '{column_name}' не найдено ни одного значения.\n\nФайл: {PRIKAZ_XLSX}"
                )
                return

            for value in values:
                value_list.insert(tk.END, value)

            debug_log("values loaded into listbox:", len(values))
            left_status_var.set(f"Колонка '{column_name}': загружено значений {len(values)}")
            value_list.focus_set()

        def on_select(_event=None):
            selection = value_list.curselection()
            debug_log("on_select() selection =", selection)
            if not selection:
                left_status_var.set("Ничего не выбрано")
                return
            selected = value_list.get(selection[0])
            debug_log("on_select() selected value =", repr(selected))
            old_var.set(selected)
            new_var.set(selected)
            left_status_var.set(f"Выбрано: {selected}")

        def on_double_click(_event=None):
            debug_log("on_double_click() called")
            on_select()
            new_entry.focus_set()
            new_entry.selection_range(0, tk.END)

        def apply_replace(*_):
            debug_log("apply_replace() called")
            old_value = old_var.get().strip()
            new_value = new_var.get().strip()
            column_name = column_var.get().strip()

            debug_log("apply_replace column =", column_name)
            debug_log("apply_replace old_value =", repr(old_value))
            debug_log("apply_replace new_value =", repr(new_value))

            if not column_name:
                messagebox.showerror("Ошибка", "Не выбрана колонка")
                return
            if not old_value:
                left_status_var.set("Сначала выбери значение в списке")
                messagebox.showerror("Ошибка", "Не выбрано исходное значение")
                return
            if not new_value:
                messagebox.showerror("Ошибка", "Введите новое значение")
                return
            if old_value == new_value:
                messagebox.showinfo("Без изменений", "Новое значение совпадает со старым")
                return

            left_status_var.set("Выполняется замена...")
            top.update_idletasks()

            try:
                changed = self.replace_prikaz_value(column_name, old_value, new_value)
                debug_log("apply_replace changed =", changed)
            except Exception as e:
                debug_log("apply_replace error =", repr(e))
                left_status_var.set(f"Ошибка: {e}")
                messagebox.showerror("Ошибка", f"Не удалось обновить приказ:\n{e}\n\nФайл: {PRIKAZ_XLSX}")
                return

            if changed == 0:
                left_status_var.set(f"Совпадения для '{old_value}' не найдены")
                messagebox.showinfo(
                    "Ничего не изменено",
                    "Совпадения не найдены. Проверь, нет ли в значении лишних пробелов, переносов строки или другого написания."
                )
                return

            refresh_values()
            old_var.set(new_value)
            new_var.set(new_value)
            left_status_var.set(f"Изменено строк: {changed}")
            refresh_summer_rows(keep_selection_name=new_value)
            messagebox.showinfo(
                "Готово",
                f"Значение '{old_value}' заменено на '{new_value}'. Всего изменено строк: {changed}.\n"
                f"Изменения применены и к первому листу, и к листу summer, если там было совпадение."
            )

        ttk.Button(left_button_frame, text="Обновить список", command=refresh_values).pack(side="left")
        ttk.Button(left_button_frame, text="Применить замену", command=apply_replace).pack(side="right")

        column_cb.bind("<<ComboboxSelected>>", refresh_values)
        value_list.bind("<<ListboxSelect>>", on_select)
        value_list.bind("<Double-Button-1>", on_double_click)
        new_entry.bind("<Return>", apply_replace)

        # -------- Правая часть: редактирование summer --------
        tk.Label(right_frame, text="Лист summer — цены:").pack(anchor="w")

        summer_table_frame = tk.Frame(right_frame)
        summer_table_frame.pack(fill="both", expand=True, pady=(0, 10))

        summer_tree = ttk.Treeview(summer_table_frame, columns=("name", "price"), show="headings", selectmode="browse")
        summer_tree.heading("name", text="Название")
        summer_tree.heading("price", text="Цена")
        summer_tree.column("name", width=380, anchor="w")
        summer_tree.column("price", width=110, anchor="center")
        summer_tree.pack(side="left", fill="both", expand=True)

        summer_scroll = tk.Scrollbar(summer_table_frame, orient="vertical", command=summer_tree.yview)
        summer_scroll.pack(side="right", fill="y")
        summer_tree.configure(yscrollcommand=summer_scroll.set)

        selected_summer_row_idx = tk.StringVar(value="")
        selected_summer_name = tk.StringVar(value="")
        summer_price_var = tk.StringVar(value="")

        tk.Label(right_frame, text="Выбранная позиция:").pack(anchor="w")
        summer_name_entry = tk.Entry(right_frame, textvariable=selected_summer_name, state="readonly")
        summer_name_entry.pack(fill="x", pady=(0, 8))

        tk.Label(right_frame, text="Новая цена:").pack(anchor="w")
        summer_price_entry = tk.Entry(right_frame, textvariable=summer_price_var)
        summer_price_entry.pack(fill="x", pady=(0, 10))

        summer_status_var = tk.StringVar(value="Лист summer загружен")
        tk.Label(right_frame, textvariable=summer_status_var, fg="#555", anchor="w", justify="left").pack(fill="x", pady=(0, 8))

        summer_button_frame = tk.Frame(right_frame)
        summer_button_frame.pack(fill="x")

        def refresh_summer_rows(keep_selection_name=None):
            debug_log("refresh_summer_rows() called")
            for item_id in summer_tree.get_children():
                summer_tree.delete(item_id)

            try:
                rows = self.get_summer_sheet_rows()
            except Exception as e:
                summer_status_var.set(f"Ошибка загрузки summer: {e}")
                messagebox.showerror("Ошибка", f"Не удалось загрузить лист summer:\n{e}")
                return

            target_item_id = None
            for row in rows:
                item_id = summer_tree.insert("", "end", values=(row["name"], row["price"]), tags=(str(row["row_idx"]),))
                if keep_selection_name and row["name"] == keep_selection_name:
                    target_item_id = item_id

            summer_status_var.set(f"Лист summer: загружено позиций {len(rows)}")

            if target_item_id:
                summer_tree.selection_set(target_item_id)
                summer_tree.focus(target_item_id)
                summer_tree.see(target_item_id)
                on_summer_select()
            else:
                selected_summer_row_idx.set("")
                selected_summer_name.set("")
                summer_price_var.set("")

        def on_summer_select(_event=None):
            selection = summer_tree.selection()
            debug_log("on_summer_select() selection =", selection)
            if not selection:
                summer_status_var.set("В summer ничего не выбрано")
                return

            item_id = selection[0]
            values = summer_tree.item(item_id, "values")
            tags = summer_tree.item(item_id, "tags")
            if not values or not tags:
                return

            row_idx = tags[0]
            name = values[0]
            price = values[1]
            selected_summer_row_idx.set(str(row_idx))
            selected_summer_name.set(str(name))
            summer_price_var.set(str(price))
            summer_status_var.set(f"Выбрано: {name}")

        def save_summer_price(*_):
            debug_log("save_summer_price() called")
            row_idx_text = selected_summer_row_idx.get().strip()
            name = selected_summer_name.get().strip()
            new_price = summer_price_var.get().strip()

            if not row_idx_text or not name:
                summer_status_var.set("Сначала выбери строку на листе summer")
                messagebox.showerror("Ошибка", "Сначала выбери позицию на листе summer")
                return
            if not new_price:
                messagebox.showerror("Ошибка", "Введите новую цену")
                return

            summer_status_var.set("Сохранение цены...")
            top.update_idletasks()

            try:
                self.update_summer_price(int(row_idx_text), new_price)
            except Exception as e:
                debug_log("save_summer_price error =", repr(e))
                summer_status_var.set(f"Ошибка: {e}")
                messagebox.showerror("Ошибка", f"Не удалось сохранить цену:\n{e}")
                return

            refresh_summer_rows(keep_selection_name=name)
            summer_price_var.set(str(new_price).strip().replace(",", "."))
            summer_status_var.set(f"Цена для '{name}' обновлена")
            messagebox.showinfo("Готово", f"Цена для '{name}' обновлена")

        ttk.Button(summer_button_frame, text="Обновить summer", command=refresh_summer_rows).pack(side="left")
        ttk.Button(summer_button_frame, text="Сохранить цену", command=save_summer_price).pack(side="right")

        summer_tree.bind("<<TreeviewSelect>>", on_summer_select)
        summer_tree.bind("<Double-Button-1>", on_summer_select)
        summer_price_entry.bind("<Return>", save_summer_price)

        refresh_values()
        refresh_summer_rows()

    def open_prikaz_for_edit(self):
        path = PRIKAZ_XLSX

        try:
            folder = os.path.dirname(path)
            if folder:
                os.makedirs(folder, exist_ok=True)

            if not os.path.exists(path):
                import pandas as pd
                pd.DataFrame(columns=["n", "doctors_name", "inspection", "analysis"]).to_excel(path, index=False)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось подготовить файл приказа для редактирования:\n{e}")
            return

        try:
            os.startfile(path)
            messagebox.showinfo(
                "Редактирование приказа",
                "Открыт файл приказа из папки проекта. Все отчёты будут использовать именно его."
            )
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")

    def report_custom(self):
        data = load_data()
        rpt = tk.Toplevel(self)
        rpt.title("Кастомный отчёт")
        rpt.resizable(False, False)
        padx, pady = 10, 5

        tk.Label(rpt, text="Период с:").grid(row=0, column=0, sticky="w", padx=padx, pady=pady)
        start_var = tk.StringVar()
        start_entry = tk.Entry(rpt, width=20, textvariable=start_var)
        start_entry.grid(row=0, column=1, sticky="w", padx=padx, pady=pady)
        tk.Button(rpt, text="📅", command=lambda: open_calendar(self, start_entry)).grid(row=0, column=2, padx=0, pady=pady)

        tk.Label(rpt, text="По:").grid(row=1, column=0, sticky="w", padx=padx, pady=pady)
        end_var = tk.StringVar()
        end_entry = tk.Entry(rpt, width=20, textvariable=end_var)
        end_entry.grid(row=1, column=1, sticky="w", padx=padx, pady=pady)
        tk.Button(rpt, text="📅", command=lambda: open_calendar(self, end_entry)).grid(row=1, column=2, padx=0, pady=pady)

        def on_start_changed(*_):
            s = start_var.get().strip()
            if is_valid_date(s):
                dt = datetime.datetime.strptime(s, "%d.%m.%Y")
                last = calendar.monthrange(dt.year, dt.month)[1]
                end_var.set(f"{last:02d}.{dt.month:02d}.{dt.year}")
        start_var.trace_add("write", on_start_changed)

        field_defs = [
            ("type", "Тип осмотра", lambda org_name, r: r.get("type", "предварительный")),
            ("organization", "Организация", lambda org_name, r: org_name),
            ("name", "ФИО", lambda org_name, r: r.get("name", "")),
            ("birthday", "Дата рождения", lambda org_name, r: r.get("birthday", "")),
            ("sex", "Пол", lambda org_name, r: r.get("sex", "")),
            ("division", "Подразделение", lambda org_name, r: r.get("division", "")),
            ("profession", "Должность", lambda org_name, r: r.get("profession", "")),
            ("factors", "Факторы", lambda org_name, r: r.get("factors", "")),
            ("typework", "Виды работ", lambda org_name, r: r.get("typework", "")),
            ("ids_date", "Дата ИДС", lambda org_name, r: r.get("ids_date", "")),
            ("diagnosis", "Диагноз", lambda org_name, r: r.get("diagnosis", "")),
            ("card_number", "№ карты", lambda org_name, r: r.get("card_number", "")),
        ]

        def collect_unique_values(getter):
            values = []
            seen = set()
            for org_name, recs in data.items():
                for r in recs:
                    value = str(getter(org_name, r) or "").strip()
                    if not value:
                        continue
                    key = value.lower()
                    if key in seen:
                        continue
                    seen.add(key)
                    values.append(value)
            values.sort(key=lambda x: x.lower())
            return [""] + values

        field_widgets = {}
        start_row = 2
        for idx, (field_key, field_label, getter) in enumerate(field_defs):
            row = start_row + idx
            enabled_var = tk.BooleanVar(value=False)
            value_var = tk.StringVar(value="")

            chk = tk.Checkbutton(rpt, text=field_label, variable=enabled_var)
            chk.grid(row=row, column=0, sticky="w", padx=padx, pady=3)

            cb = Combobox(rpt, textvariable=value_var, values=collect_unique_values(getter), width=42, state="disabled")
            cb.grid(row=row, column=1, columnspan=2, sticky="ew", padx=padx, pady=3)

            def make_toggle(var=enabled_var, combo=cb):
                def toggle_state():
                    combo.config(state="readonly" if var.get() else "disabled")
                    if not var.get():
                        combo.set("")
                return toggle_state

            enabled_var.trace_add("write", lambda *_args, toggle=make_toggle(): toggle())

            field_widgets[field_key] = {
                "label": field_label,
                "enabled_var": enabled_var,
                "value_var": value_var,
                "getter": getter,
            }

        def make_report_custom():
            start = start_var.get().strip()
            end = end_var.get().strip()

            if start and not is_valid_date(start):
                messagebox.showerror("Ошибка даты", "Дата начала должна быть в формате ДД.ММ.ГГГГ")
                return
            if end and not is_valid_date(end):
                messagebox.showerror("Ошибка даты", "Дата конца должна быть в формате ДД.ММ.ГГГГ")
                return

            d0 = datetime.datetime.strptime(start, "%d.%m.%Y") if start else None
            d1 = datetime.datetime.strptime(end, "%d.%m.%Y") if end else None
            if d0 is not None and d1 is not None and d1 < d0:
                messagebox.showerror("Ошибка", "Конечная дата меньше начальной")
                return

            selected_fields = []
            active_filters = {}
            for field_key, cfg in field_widgets.items():
                if cfg["enabled_var"].get():
                    selected_fields.append(field_key)
                    selected_value = cfg["value_var"].get().strip()
                    # Если поле включено, но значение не выбрано,
                    # то поле просто попадёт в отчёт без фильтрации.
                    if selected_value:
                        active_filters[field_key] = selected_value

            if not selected_fields:
                messagebox.showerror("Ошибка", "Выбери хотя бы одно поле для отчёта")
                return

            rows = []
            for org_name, recs in data.items():
                for r in recs:
                    ids = (r.get("ids_date") or "").strip()
                    if not ids:
                        continue
                    try:
                        d_ids = datetime.datetime.strptime(ids, "%d.%m.%Y")
                    except ValueError:
                        continue
                    if d0 is not None and d_ids < d0:
                        continue
                    if d1 is not None and d_ids > d1:
                        continue

                    passed = True
                    for field_key, selected_value in active_filters.items():
                        current_value = str(field_widgets[field_key]["getter"](org_name, r) or "").strip()
                        if current_value != selected_value:
                            passed = False
                            break
                    if not passed:
                        continue

                    row = {}
                    for field_key in selected_fields:
                        cfg = field_widgets[field_key]
                        row[cfg["label"]] = str(cfg["getter"](org_name, r) or "").strip()
                    rows.append(row)

            if not rows:
                messagebox.showinfo("Пустой отчет", "Нет записей за выбранный период и фильтры.")
                return

            df = pd.DataFrame(rows)
            save_dir = self.settings.get("save_dir", os.getcwd())
            period_start = start if start else "все"
            period_end = end if end else "все"
            fname = sanitize_filename(f"Кастомный_отчет_{period_start}_{period_end}.xlsx")
            save_path = os.path.join(save_dir, fname)

            try:
                from openpyxl.utils import get_column_letter
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Report")
                    sheet = writer.sheets["Report"]
                    for idx, col in enumerate(df.columns, start=1):
                        width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                        sheet.column_dimensions[get_column_letter(idx)].width = min(width, 60)
                show_notification(self, f"Отчет сохранён: {save_path}")
                rpt.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка записи", f"Не удалось сохранить файл:\n{e}")

        tk.Button(
            rpt,
            text="Сформировать",
            command=make_report_custom,
            bg="#4CAF50",
            fg="white"
        ).grid(row=start_row + len(field_defs), column=0, columnspan=3, pady=(12, 10), padx=padx, sticky="ew")

        rpt.grid_columnconfigure(1, weight=1)
        rpt.update_idletasks()
        w, h = rpt.winfo_width(), rpt.winfo_height()
        sw, sh = rpt.winfo_screenwidth(), rpt.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 2
        rpt.geometry(f"{max(w, 620)}x{h}+{x}+{y}")

    # ============ ОТЧЁТ ПО ОРГАНИЗАЦИИ ============
    def report_by_organization(self):
        data = load_data()
        rpt = tk.Toplevel(self)
        rpt.title("Отчет по организации")
        rpt.resizable(False, False)
        padx, pady = 10, 5
        org_var_report = tk.StringVar()
        tk.Label(rpt, text="Организация:").grid(row=0, column=0, sticky="w", padx=padx, pady=pady)
        org_list = sorted(data.keys())
        org_cb = Combobox(
            rpt,
            values=org_list,
            textvariable=org_var_report,
            width=40,
            state="readonly"
        )
        org_cb.grid(row=0, column=1, padx=padx, pady=pady)
        org_var_report.set("")
        tk.Label(rpt, text="Период с:").grid(row=1, column=0, sticky="w", padx=padx, pady=pady)
        start_var = tk.StringVar()
        start_entry = tk.Entry(rpt, width=20, textvariable=start_var)
        start_entry.grid(row=1, column=1, sticky="w", padx=padx, pady=pady)
        tk.Button(rpt, text="📅", command=lambda: open_calendar(self, start_entry)).grid(row=1, column=2, padx=0, pady=pady)
        tk.Label(rpt, text="По:").grid(row=2, column=0, sticky="w", padx=padx, pady=pady)
        end_var = tk.StringVar()
        end_entry = tk.Entry(rpt, width=20, textvariable=end_var)
        end_entry.grid(row=2, column=1, sticky="w", padx=padx, pady=pady)
        tk.Button(rpt, text="📅", command=lambda: open_calendar(self, end_entry)).grid(row=2, column=2, padx=0, pady=pady)

        def on_start_changed(*_):
            s = start_var.get().strip()
            if is_valid_date(s):
                dt = datetime.datetime.strptime(s, "%d.%m.%Y")
                last = calendar.monthrange(dt.year, dt.month)[1]
                end_var.set(f"{last:02d}.{dt.month:02d}.{dt.year}")
        start_var.trace_add("write", on_start_changed)

        def make_report():
            org_sel = org_var_report.get().strip()
            start = start_entry.get().strip()
            end = end_entry.get().strip()
            if not org_sel:
                messagebox.showerror("Ошибка ввода", "Выберите организацию")
                return
            if not is_valid_date(start) or not is_valid_date(end):
                messagebox.showerror("Ошибка даты", "Даты в формате ДД.ММ.ГГГГ")
                return
            d0 = datetime.datetime.strptime(start, "%d.%m.%Y")
            d1 = datetime.datetime.strptime(end, "%d.%m.%Y")
            if d1 < d0:
                messagebox.showerror("Ошибка", "Конечная дата меньше начальной")
                return
            rows = []
            for r in data.get(org_sel, []):
                ids = r.get("ids_date", "").strip()
                if not ids:
                    continue
                try:
                    d_ids = datetime.datetime.strptime(ids, "%d.%m.%Y")
                except ValueError:
                    continue
                if d0 <= d_ids <= d1:
                    rows.append({
                        "Тип осмотра": r.get("type", "предварительный"),
                        "Организация": org_sel,
                        "ФИО": r["name"],
                        "Дата рожд.": r["birthday"],
                        "Пол": r["sex"],
                        "Подразделение": r["division"],
                        "Должность": r["profession"],
                        "Факторы": r["factors"],
                        "Виды работ": r["typework"],
                        "Дата ИДС": ids,
                        "Диагноз": r.get("diagnosis", ""),
                        "№ карты": r.get("card_number", "")
                    })
            if not rows:
                messagebox.showinfo("Пустой отчет", "Нет записей за выбранный период.")
                return
            df = pd.DataFrame(rows)
            save_dir = self.settings.get("save_dir", os.getcwd())
            fname = sanitize_filename(f"Отчет_{org_sel}_{start}_{end}.xlsx")
            save_path = os.path.join(save_dir, fname)
            try:
                from openpyxl.utils import get_column_letter
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Report")
                    sheet = writer.sheets["Report"]
                    for idx, col in enumerate(df.columns, start=1):
                        width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                        sheet.column_dimensions[get_column_letter(idx)].width = width
                show_notification(self, f"Отчет сохранён: {save_path}")
                rpt.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка записи", f"Не удалось сохранить файл:\n{e}")

        tk.Button(
            rpt,
            text="Сформировать",
            command=make_report,
            bg="#4CAF50",
            fg="white"
        ).grid(row=3, column=0, columnspan=3, pady=(10, 10), padx=padx, sticky="ew")
        rpt.grid_columnconfigure(1, weight=1)
        rpt.update_idletasks()
        w, h = rpt.winfo_width(), rpt.winfo_height()
        sw, sh = rpt.winfo_screenwidth(), rpt.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 2
        rpt.geometry(f"{w}x{h}+{x}+{y}")

    # ============ ОТЧЁТ ПО МЕСЯЦУ ============
    def report_by_month(self):
        data = load_data()
        rpt = tk.Toplevel(self)
        rpt.title("Отчет по дате ИДС")
        rpt.resizable(False, False)
        padx, pady = 10, 5
        tk.Label(rpt, text="Период с:").grid(row=0, column=0, sticky="w", padx=padx, pady=pady)
        start_var = tk.StringVar()
        start_entry = tk.Entry(rpt, width=20, textvariable=start_var)
        start_entry.grid(row=0, column=1, sticky="w", padx=padx, pady=pady)
        tk.Button(rpt, text="📅", command=lambda: open_calendar(self, start_entry)).grid(row=0, column=2, padx=0, pady=pady)
        tk.Label(rpt, text="По:").grid(row=1, column=0, sticky="w", padx=padx, pady=pady)
        end_var = tk.StringVar()
        end_entry = tk.Entry(rpt, width=20, textvariable=end_var)
        end_entry.grid(row=1, column=1, sticky="w", padx=padx, pady=pady)
        tk.Button(rpt, text="📅", command=lambda: open_calendar(self, end_entry)).grid(row=1, column=2, padx=0, pady=pady)
        def on_start_changed(*_):
            s = start_var.get().strip()
            if is_valid_date(s):
                dt = datetime.datetime.strptime(s, "%d.%m.%Y")
                last = calendar.monthrange(dt.year, dt.month)[1]
                end_var.set(f"{last:02d}.{dt.month:02d}.{dt.year}")
        start_var.trace_add("write", on_start_changed)
        def make_report_month():
            start = start_var.get().strip()
            end = end_var.get().strip()
            d0 = datetime.datetime.strptime(start, "%d.%m.%Y")
            d1 = datetime.datetime.strptime(end, "%d.%m.%Y")
            rows = []
            for org_name, recs in data.items():
                for r in recs:
                    ids = r.get("ids_date", "").strip()
                    if not ids:
                        continue
                    try:
                        d_ids = datetime.datetime.strptime(ids, "%d.%m.%Y")
                    except ValueError:
                        continue
                    if d0 <= d_ids <= d1:
                        rows.append({
                            "Тип осмотра": r.get("type", "предварительный"),
                            "Организация": org_name,
                            "ФИО": r["name"],
                            "Дата рожд.": r["birthday"],
                            "Пол": r["sex"],
                            "Подразделение": r["division"],
                            "Должность": r["profession"],
                            "Факторы": r["factors"],
                            "Виды работ": r["typework"],
                            "Дата ИДС": ids,
                            "Диагноз": r.get("diagnosis", ""),
                            "№ карты": r.get("card_number", "")
                        })
            if not rows:
                messagebox.showinfo("Пустой отчет", "Нет записей за выбранный период.")
                return
            df = pd.DataFrame(rows)
            save_dir = self.settings.get("save_dir", os.getcwd())
            fname = sanitize_filename(f"Отчет_по_месяцу_{start}_{end}.xlsx")
            save_path = os.path.join(save_dir, fname)
            try:
                from openpyxl.utils import get_column_letter
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Report")
                    sheet = writer.sheets["Report"]
                    for idx, col in enumerate(df.columns, start=1):
                        width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                        sheet.column_dimensions[get_column_letter(idx)].width = width
                show_notification(self, f"Отчет сохранён: {save_path}")
                rpt.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка записи", f"Не удалось сохранить файл:\n{e}")
        tk.Button(
            rpt,
            text="Сформировать",
            command=make_report_month,
            bg="#4CAF50",
            fg="white"
        ).grid(row=2, column=0, columnspan=3, pady=(10, 10), padx=padx, sticky="ew")
        rpt.grid_columnconfigure(1, weight=1)
        rpt.update_idletasks()
        w, h = rpt.winfo_width(), rpt.winfo_height()
        sw, sh = rpt.winfo_screenwidth(), rpt.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 2
        rpt.geometry(f"{w}x{h}+{x}+{y}")




if __name__ == "__main__":
    setup_logging()
    debug_log("program entry")
    app = MainApp()
    debug_log("MainApp created")
    app.mainloop()
    debug_log("mainloop finished")
