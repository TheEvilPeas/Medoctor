import tkinter as tk
import tkinter.ttk as ttk
import pandas as pd
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys, os, json


def settings_path():
    path = resource_path("res/settings.json")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    return path

def resource_path(rel_path: str) -> str:
    """
    Возвращает путь к ресурсу и в dev-режиме, и внутри PyInstaller.
    rel_path: относительный путь внутри проекта (например 'conclusion_form/res/template.docx')
    """
    if hasattr(sys, '_MEIPASS'):
        base = sys._MEIPASS  # временная папка PyInstaller
    else:
        base = os.path.abspath(".")
    return os.path.join(base, rel_path)

SETTINGS_PATH = settings_path()
TEMPLATE_PATH = resource_path("conclusion_form/res/template.docx")
XML_PATH      = resource_path("conclusion_form/res/data.xml")

CALENDAR_PNG  = resource_path("conclusion_form/res/calendar.png")
PRIKAZ_XLSX   = resource_path("search_form/input/prikaz29n.xlsx")

def get_prikaz_read_path():
    return PRIKAZ_XLSX

def open_prikaz_for_edit():
    path = PRIKAZ_XLSX
    try:
        folder = os.path.dirname(path)
        if folder:
            os.makedirs(folder, exist_ok=True)
        if not os.path.exists(path):
            pd.DataFrame(columns=["n", "doctors_name", "inspection", "analysis"]).to_excel(path, index=False)

        if sys.platform == "darwin":
            import subprocess
            subprocess.Popen(["open", path])
        elif os.name == "nt":
            os.startfile(path)
        else:
            import subprocess
            subprocess.Popen(["xdg-open", path])
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось подготовить/открыть файл приказа:\n{e}")


class SearchForm(tk.Frame):
    def __init__(self, parent, main_app=None):
        super().__init__(parent)
        self.main_app = main_app
        self.entries = []
        self.gender_m_var = tk.BooleanVar()
        self.gender_f_var = tk.BooleanVar()
        self.age_over40_var = tk.BooleanVar()
        self.age_under40_var = tk.BooleanVar()
        # --- Загрузка данных ---
        self.df = pd.read_excel(get_prikaz_read_path())
        self.df['n'] = self.df['n'].astype(str).str.replace(',', '.')
        self.build_ui()
        self.last_save_path = None

    def render_preview(self, df: pd.DataFrame):
        # очистка старых колонок/строк
        for col in self.preview_tree["columns"]:
            self.preview_tree.heading(col, text="")
            self.preview_tree.column(col, width=0)
        self.preview_tree.delete(*self.preview_tree.get_children())

        # настройка колонок по df
        cols = list(df.columns.astype(str))
        self.preview_tree["columns"] = cols
        for col in cols:
            self.preview_tree.heading(col, text=col)
            # примерная ширина по длине заголовка
            self.preview_tree.column(col, width=max(80, min(300, len(col) * 10)))

        # вставка строк
        for _, row in df.iterrows():
            values = ["" if pd.isna(v) else str(v) for v in row.tolist()]
            self.preview_tree.insert("", "end", values=values)

    def _toggle_gender_m(self):
        if self.gender_m_var.get():
            self.gender_f_var.set(False)

    def _toggle_gender_f(self):
        if self.gender_f_var.get():
            self.gender_m_var.set(False)

    def _toggle_age_over40(self):
        if self.age_over40_var.get():
            self.age_under40_var.set(False)

    def _toggle_age_under40(self):
        if self.age_under40_var.get():
            self.age_over40_var.set(False)

    def build_ui(self):
        # Чекбоксы
        checkbox_frame = tk.Frame(self)
        checkbox_frame.pack(pady=10)
        tk.Label(checkbox_frame, text="Пол:").grid(row=0, column=0, sticky='w')
        tk.Checkbutton(checkbox_frame, text="М", variable=self.gender_m_var, command=self._toggle_gender_m).grid(row=0,
                                                                                                                 column=1,
                                                                                                                 sticky='w')
        tk.Checkbutton(checkbox_frame, text="Ж", variable=self.gender_f_var, command=self._toggle_gender_f).grid(row=0,
                                                                                                                 column=2,
                                                                                                                 sticky='w')

        tk.Label(checkbox_frame, text="Возраст:").grid(row=1, column=0, sticky='w')
        tk.Checkbutton(checkbox_frame, text=">40", variable=self.age_over40_var, command=self._toggle_age_over40).grid(
            row=1, column=1, sticky='w')
        tk.Checkbutton(checkbox_frame, text="<40", variable=self.age_under40_var,
                       command=self._toggle_age_under40).grid(row=1, column=2, sticky='w')

        # Кнопки для добавления/удаления полей
        button_frame = tk.Frame(self)
        button_frame.pack()
        tk.Button(button_frame, text="Добавить поле", command=self.add_entry).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Удалить поле", command=self.remove_entry).pack(side=tk.LEFT, padx=5)

        # Кнопка поиска
        tk.Button(self, text="Поиск", command=self.search_items).pack(pady=10)
        tk.Button(self, text="Печать…", command=self.print_results).pack(pady=(0, 10))


        # Фрейм для ввода номеров пунктов
        # --- Прокручиваемый список полей (до 5 видимых, по центру) ---
        entries_container = tk.Frame(self)
        entries_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.entries_canvas = tk.Canvas(entries_container, highlightthickness=0)
        vsb = tk.Scrollbar(entries_container, orient="vertical", command=self.entries_canvas.yview)
        self.entries_canvas.configure(yscrollcommand=vsb.set)

        self.entries_canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # внутренние фреймы
        self.entries_inner = tk.Frame(self.entries_canvas)
        # окно внутри канваса — СОХРАНЯЕМ ID!
        self.entries_window = self.entries_canvas.create_window((0, 0), window=self.entries_inner, anchor="n")

        # центрирующий фрейм
        self.center_frame = tk.Frame(self.entries_inner)
        self.center_frame.pack()

        def _on_inner_configure(event):
            # обновляем scrollregion
            self.entries_canvas.configure(scrollregion=self.entries_canvas.bbox("all"))

        self.entries_inner.bind("<Configure>", _on_inner_configure)

        def _on_canvas_configure(event):
            # растягиваем внутреннее окно по ширине канваса
            self.entries_canvas.itemconfig(self.entries_window, width=event.width)

        self.entries_canvas.bind("<Configure>", _on_canvas_configure)

        # фиксируем высоту примерно на 5 строк
        ROW_H = 28
        self.entries_canvas.config(height=ROW_H * 5)

        # прокрутка колесом (Windows/*nix)
        def _on_mousewheel(event):
            try:
                step = -1 if event.delta > 0 else 1
                # если виджет уже уничтожен — не скроллим
                if not self.entries_canvas.winfo_exists():
                    return
                self.entries_canvas.yview_scroll(step, "units")
            except tk.TclError:
                pass

        # стало — привязки только к канвасу
        self.entries_canvas.bind("<MouseWheel>", _on_mousewheel)  # Windows / macOS
        self.entries_canvas.bind("<Button-4>", lambda e: self.entries_canvas.yview_scroll(-1, "units"))  # *nix up
        self.entries_canvas.bind("<Button-5>", lambda e: self.entries_canvas.yview_scroll(1, "units"))  # *nix down

        # первое поле
        self.add_entry()

        # --- Предпросмотр (как листы в Excel) ---
        self.preview_area = tk.Frame(self)
        self.preview_area.pack(fill="both", expand=True, padx=10, pady=(5, 0))

        self.preview_notebook = ttk.Notebook(self.preview_area)
        self.preview_notebook.pack(fill="both", expand=True)

        # Лист "Результаты"
        tab_results = ttk.Frame(self.preview_notebook)
        self.preview_notebook.add(tab_results, text="Результаты")
        rs_vsb = tk.Scrollbar(tab_results, orient="vertical")
        rs_hsb = tk.Scrollbar(tab_results, orient="horizontal")
        self.preview_main_tree = ttk.Treeview(tab_results, show="headings",
                                              yscrollcommand=rs_vsb.set, xscrollcommand=rs_hsb.set)
        rs_vsb.config(command=self.preview_main_tree.yview)
        rs_hsb.config(command=self.preview_main_tree.xview)
        self.preview_main_tree.grid(row=0, column=0, sticky="nsew")
        rs_vsb.grid(row=0, column=1, sticky="ns")
        rs_hsb.grid(row=1, column=0, sticky="ew")
        tab_results.rowconfigure(0, weight=1)
        tab_results.columnconfigure(0, weight=1)

        # Лист "Summer"
        tab_summer = ttk.Frame(self.preview_notebook)
        self.preview_notebook.add(tab_summer, text="Summer")
        sm_vsb = tk.Scrollbar(tab_summer, orient="vertical")
        sm_hsb = tk.Scrollbar(tab_summer, orient="horizontal")
        self.preview_summer_tree = ttk.Treeview(tab_summer, show="headings",
                                                yscrollcommand=sm_vsb.set, xscrollcommand=sm_hsb.set)
        sm_vsb.config(command=self.preview_summer_tree.yview)
        sm_hsb.config(command=self.preview_summer_tree.xview)
        self.preview_summer_tree.grid(row=0, column=0, sticky="nsew")
        sm_vsb.grid(row=0, column=1, sticky="ns")
        sm_hsb.grid(row=1, column=0, sticky="ew")
        tab_summer.rowconfigure(0, weight=1)
        tab_summer.columnconfigure(0, weight=1)

        self.summer_total_var = tk.StringVar(value="Итоговая сумма: 0")
        self.summer_total_bar = tk.Frame(self)
        self.summer_total_bar.pack(side="bottom", fill="x", padx=10, pady=(6, 10))

        self.summer_total_label = tk.Label(
            self.summer_total_bar,
            textvariable=self.summer_total_var,
            anchor="w",
            relief="groove",
            bd=1,
            padx=8,
            pady=4
        )
        self.summer_total_label.pack(fill="x")

        # Лист "Summer"
        # (duplicate block removed)

    def print_results(self):
        if not self.last_save_path or not os.path.exists(self.last_save_path):
            messagebox.showerror("Печать", "Сначала сформируйте файл (нажмите «Поиск»).")
            return

        if os.name != "nt":
            messagebox.showerror("Печать", "Печать через Excel доступна только в Windows.")
            return

        try:
            import win32com.client as win32
        except ImportError:
            messagebox.showerror(
                "Печать",
                "В этой сборке отсутствует модуль 'pywin32'.\n\n"
                "Чтобы печать работала на всех ПК без установки вручную,\n"
                "нужно установить 'pywin32' на Windows-компьютере перед сборкой программы\n"
                "и пересобрать приложение."
            )
            return

        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = True
        wb = excel.Workbooks.Open(self.last_save_path)
        excel.Application.Dialogs(8).Show()

    def render_to_tree(self, tree: ttk.Treeview, df: pd.DataFrame):
        # очистка
        for col in tree["columns"]:
            tree.heading(col, text="")
            tree.column(col, width=0)
        tree.delete(*tree.get_children())

        if df is None or df.empty:
            tree["columns"] = []
            return

        cols = [str(c) for c in df.columns]
        tree["columns"] = cols
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=max(80, min(300, len(col) * 10)))

        for _, row in df.iterrows():
            values = ["" if pd.isna(v) else str(v) for v in row.tolist()]
            tree.insert("", "end", values=values)

    def add_entry(self):
        entry = tk.Entry(self.center_frame)
        entry.pack(pady=5)
        self.entries.append(entry)

    def remove_entry(self):
        if self.entries:
            entry = self.entries.pop()
            entry.destroy()

    def search_items(self):
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        import os

        item_numbers = []
        for entry in self.entries:
            if entry.get():
                item_numbers.extend([s.strip() for s in entry.get().split(',') if s.strip()])

        # --- ДОБАВЛЯЕМ специальные пункты по выбранным чекбоксам ---
        if self.gender_m_var.get() and self.age_under40_var.get():
            item_numbers.append('0.11')
        if self.gender_m_var.get() and self.age_over40_var.get():
            item_numbers.append('0.12')
        if self.gender_f_var.get() and self.age_under40_var.get():
            item_numbers.append('0.21')
        if self.gender_f_var.get() and self.age_over40_var.get():
            item_numbers.append('0.22')

        filtered_data = self.df[self.df['n'].isin(item_numbers)]

        # Врачи
        doctors_split = filtered_data['doctors_name'].str.split(',')
        unique_doctors = list(
            set([doctor.strip() for sublist in doctors_split.dropna() for doctor in sublist if doctor.strip()]))

        # Обследования
        inspection_split = filtered_data['inspection'].apply(
            lambda x: str(x).split(',') if pd.notna(x) else [])
        unique_inspections = list(set(
            [inspection.strip() for sublist in inspection_split for inspection in sublist if inspection.strip()]))

        # Анализы
        analysis_split = filtered_data['analysis'].apply(
            lambda x: str(x).split(',') if pd.notna(x) else [])
        unique_analysis = list(set(
            [analysis.strip() for sublist in analysis_split for analysis in sublist if analysis.strip()]))

        max_len = max(len(unique_doctors), len(unique_analysis), len(unique_inspections))
        unique_doctors += [''] * (max_len - len(unique_doctors))
        unique_inspections += [''] * (max_len - len(unique_inspections))
        unique_analysis += [''] * (max_len - len(unique_analysis))

        data = {'Врачи': unique_doctors, 'Обследования': unique_inspections, 'Анализы': unique_analysis}
        df_unique = pd.DataFrame(data)

        # Сохраняем файл в папку из настроек
        save_dir = self.main_app.settings.get("save_dir", os.getcwd())
        save_path = os.path.join(save_dir, 'results.xlsx')
        self.last_save_path = save_path

        # Пишем основной лист (Результаты) с защитой от PermissionError
        try:
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                df_unique.to_excel(writer, index=False, sheet_name="Sheet1")
        except PermissionError:
            messagebox.showerror(
                "Ошибка",
                "Невозможно провести поиск: нет доступа к файлу результатов.\n"
                "Возможно, файл 'results.xlsx' открыт в другой программе."
            )
            return

        # Предпросмотр "Результаты"
        self.render_to_tree(self.preview_main_tree, df_unique)

        # (Summer tab is preview-only; do not add to Excel file)

        # Автоматическая подгонка ширины столбцов первого листа
        try:
            results_workbook = load_workbook(save_path)
            results_sheet = results_workbook["Sheet1"]
            for column in results_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                results_sheet.column_dimensions[column_letter].width = adjusted_width
            results_workbook.save(save_path)
        except PermissionError:
            messagebox.showerror(
                "Ошибка",
                "Невозможно обновить файл результатов (автоширина).\n"
                "Возможно, файл 'results.xlsx' открыт в другой программе."
            )
            return

        # --- Предпросмотр "Summer" только в программе, без записи в results.xlsx ---
        try:
            if os.path.exists(PRIKAZ_XLSX):
                prikaz_workbook = load_workbook(PRIKAZ_XLSX, data_only=True)
                if len(prikaz_workbook.sheetnames) > 1:
                    summer_sheet = prikaz_workbook[prikaz_workbook.sheetnames[1]]
                    summer_rows = list(summer_sheet.iter_rows(values_only=True))

                    if summer_rows:
                        summer_df = pd.DataFrame(summer_rows)
                        if not summer_df.empty and summer_df.iloc[0].notna().any():
                            summer_df.columns = summer_df.iloc[0].astype(str)
                            summer_df = summer_df.iloc[1:].reset_index(drop=True)

                        cols = [c for c in ['Врачи', 'Обследования', 'Анализы'] if c in df_unique.columns]
                        ref_values = pd.concat([df_unique[c].dropna().astype(str).str.strip() for c in cols], ignore_index=True)
                        ref_set = set(v for v in ref_values if v)

                        sum_value = 0
                        if len(summer_df.columns):
                            check_col = summer_df.columns[0]
                            summer_df['Соответствие'] = (
                                summer_df[check_col].astype(str).str.strip()
                                .apply(lambda v: '+' if v and v in ref_set else '')
                            )

                            if len(summer_df.columns) >= 2:
                                values_col = summer_df.columns[1]
                                matched_mask = summer_df['Соответствие'].astype(str).str.strip() == '+'
                                sum_value = pd.to_numeric(
                                    summer_df.loc[matched_mask, values_col],
                                    errors='coerce'
                                ).fillna(0).sum()

                        self.render_to_tree(self.preview_summer_tree, summer_df)
                        self.summer_total_var.set(f"Итоговая сумма: {sum_value}")
                    else:
                        self.render_to_tree(self.preview_summer_tree, pd.DataFrame())
                        self.summer_total_var.set("Итоговая сумма: 0")
                else:
                    self.render_to_tree(self.preview_summer_tree, pd.DataFrame())
                    self.summer_total_var.set("Итоговая сумма: 0")
            else:
                self.render_to_tree(self.preview_summer_tree, pd.DataFrame())
                self.summer_total_var.set("Итоговая сумма: 0")
        except Exception as e:
            messagebox.showerror("Ошибка Summer", str(e))
            self.render_to_tree(self.preview_summer_tree, pd.DataFrame())
            self.summer_total_var.set("Итоговая сумма: 0")






